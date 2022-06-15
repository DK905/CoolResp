r"""Сохранение БД расписания группы в Excel-документ с нативной вёрсткой.

"""

import cr_component.parser.additional as cr_add
import cr_component.parser.defaults as cr_def
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Border, Font, GradientFill, PatternFill, Side


""" Константные стили и названия для оформления """
# Стиль шапки
ST_TITLE = NamedStyle(name='Шапка')
ST_TITLE.font = Font(name='Book Antiqua', size=14)
ST_TITLE.alignment = Alignment(horizontal='center', vertical='center')
ST_TITLE.border = Border(left=Side(border_style='thick'), right=Side(border_style='thick'),
                         top=Side(border_style='thick'), bottom=Side(border_style='thick'))

# Стиль дней
ST_DAYS = NamedStyle(name='Дни')
ST_DAYS.font = Font(name='Bookman Old Style', size=14, bold=True)
ST_DAYS.alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
ST_DAYS.border = Border(left=Side(border_style='thick'), right=Side(border_style='thick'),
                        top=Side(border_style='thick'), bottom=Side(border_style='thick'))

# Стиль базовой ячейки
ST_COMMON = NamedStyle(name='Базовая ячейка')
ST_COMMON.font = Font(name='Plantagenet Cherokee', size=14)
ST_COMMON.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ST_COMMON.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                          top=Side(border_style='thin'), bottom=Side(border_style='thin'))

# Стиль для номеров пары и времени
ST_INFO = NamedStyle(name='Инфополе')
ST_INFO.font = Font(name='Plantagenet Cherokee', size=14)
ST_INFO.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ST_INFO.border = Border(left=Side(border_style='thick'), right=Side(border_style='thick'),
                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))

# Стиль пустой ячейки
ST_NULL = NamedStyle(name='Круговерть пустоты')
ST_NULL.font = Font(name='Plantagenet Cherokee', size=14)
ST_NULL.border = Border(left=Side(border_style=None), right=Side(border_style=None),
                        top=Side(border_style=None), bottom=Side(border_style=None))
ST_NULL_COLOR = '0000ff27'
ST_NULL.fill = PatternFill(patternType='lightDown', start_color=ST_NULL_COLOR)

# Названия столбцов
parse_title = ['Дни', '№ пары', 'Время', 'Ауд', 'Преподаватель']


def time_resp(df:  'БД расписания'
              ) -> 'Датафрейм границ учебных недель: ПН - СБ - строчная запись диапазона':

    """ Функция определения периода рабочих дней недель (для дней в шапке) """

    # Первый день расписания
    date_min = df['date_pair'].dropna().min()
    # Последний день расписания
    date_max = df['date_pair'].dropna().max()

    # Если расписание начинается позже понедельника
    if date_min.dayofweek != 0:
        # Найти ближайший (слева) понедельник
        date_min -= date_min.dayofweek * pd.tseries.offsets.Day()

    # Если расписание заканчивается раньше субботы
    if date_max.dayofweek < 5:
        # Найти ближайшую (справа) субботу
        date_max += (5 - date_max.dayofweek) * pd.tseries.offsets.Day()

    # Серия понедельников
    date_monday = pd.Series(pd.date_range(date_min, date_max, freq="W-MON"))
    # Серия суббот
    date_saturday = pd.Series(pd.date_range(date_min, date_max, freq="W-SAT"))

    # Датафрейм периодов рабочих недель
    work_weeks = pd.DataFrame({'monday': date_monday, 'saturday': date_saturday})

    # Функция форматной записи учебной недели
    date_format = lambda date_inf: f"{date_inf.strftime('%d')} {cr_def.MONTHS.loc[date_inf.month, 'abbr_name']}"

    # Отдельный столбец с форматной записью периода
    work_weeks['diap_string'] = work_weeks.apply(lambda row: ' - '.join([date_format(row['monday']),
                                                                         date_format(row['saturday'])]),
                                                 axis=1)

    # Отдельный столбец для хранения названия листа, на котором расположена учебная неделя
    work_weeks['ind_sheet'] = ''

    # Отдельный столбец для хранения индекса столбца учебной недели
    work_weeks['ind_col'] = 1

    # Возврат списка границ учебных недель
    return work_weeks


def sheet_and_headers(wb:  'Объект книги EXCEL из openpyxl',
                      tt:  'Список границ учебных недель'
                      ) -> 'Книга с заполненной основой: шапки, листы и базовые стили':

    """ Функция заполнения листов таблицы и их шапок """

    # Заполнение листов таблицы и их шапок
    for d_ind, week in tt.iterrows():
        # Учебный месяц изменяется при определённом наборе условий
        if (  # Если запись первая
            not d_ind or
            # Или запись не первая
            (d_ind and
             # Где месяц отличается у понедельника текущей и предыдущей недели
             (tt.loc[d_ind-1]['monday'].month != week['monday'].month or
              tt.loc[d_ind-1]['saturday'].month != week['saturday'].month) and
             # И понедельник/суббота приходятся на первые числа месяца
             (int(week['monday'].day) in range(1, 8) or
              int(week['saturday'].day) in range(1, 8)) and
             # И месяца ещё не было в списке листов
             cr_def.MONTHS.loc[week['monday'].month, 'full_name'] not in wb.sheetnames)):

            # Если в месяце была одна неделя (при смене месяца), то нужно просто объединить месяцы
            if not wb.sheetnames[0] == 'Sheet' and ws.max_column < 5:
                # Переименовать текущий месяц
                ws.title = cr_def.MONTHS.loc[week['monday'].month, 'full_name']

            # Если в следующем месяце будет всего одна неделя (на случай чумы или сессии), то ничего не менять
            elif (  # Если не выполняется групповое условие, где текущая запись - последняя
                  not(d_ind+1 == tt.shape[0] or
                      # Или запись не последняя
                      d_ind and
                      # Но месяц понедельника предыдущей, текущей и следующей записей не совпадает
                      tt.loc[d_ind-1, 'monday'].month != week['monday'].month != tt.loc[d_ind+1, 'monday'].month)):

                # Если добавляется первый лист, то можно просто переименовать стандартный
                if wb.sheetnames[0] == 'Sheet':
                    # Переключиться на актуальный лист таблицы
                    ws = wb.active
                    # Переименовать актуальный лист
                    ws.title = cr_def.MONTHS.loc[week['monday'].month, 'full_name']
                # Если обычный новый учебный месяц, то создать новый лист
                else:
                    # Создать лист
                    wb.create_sheet(cr_def.MONTHS.loc[week['monday'].month, 'full_name'])

                # Сделать актуальным листом последний лист
                ws = wb.worksheets[-1]
                # Заполнение первых трёх столбцов шапки
                _ = ws.cell(column=ws.max_column,   row=1, value=parse_title[0]).style = ST_TITLE
                _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[1]).style = ST_TITLE
                _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[2]).style = ST_TITLE

        # Запись учебной недели в красивом формате
        _ = ws.cell(column=ws.max_column+1, row=1, value=week['diap_string']).style = ST_TITLE

        # Если неделя была последней
        if d_ind == tt.shape[0]-1:
            # Предыдущий лист - текущий лист
            pred = wb.worksheets[-1]
        # Если произошла смена листа
        elif ws.max_column == 4 and len(wb.worksheets) > 1:
            # Предыдущий лист - пред-текущий лист
            pred = wb.worksheets[-2]
        # Если всё ок, то пропустить заполнение последних стилей шапки
        else:
            continue

        # Эти стили нужно применять только когда лист уже заполнен
        # Применить стиль к кабинетам
        _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[3]).style = ST_TITLE
        # Применить стиль к преподам
        _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[4]).style = ST_TITLE

    # Отдельная пробежка по листам для занесения данных об учебных неделях: не в первом цикле из-за глюков openpyxl
    ind = 0
    for ws in wb:
        for i in range(4, ws.max_column-1):
            # Занесение данных об учебной неделе в изначальные данные
            # Сохранить название листа с учебной неделей
            tt.loc[ind, 'ind_sheet'] = ws.title
            # Сохранить номер столбца учебной недели
            tt.loc[ind, 'ind_col'] = i
            ind += 1

    # Вернуть книгу с готовым шаблоном и обновление списка учебных недель
    return wb


def fill_base(wb: 'Объект книги EXCEL из openpyxl',
              df: 'База парсинга',
              tt: 'Список границ учебных недель'
              ) -> 'Заполненная, но не отформатированная, книга с расписанием':

    """ Функция заполнения таблицы расписанием """

    # Сортировка датафрейма в порядке "день - номер - дата"
    df = df.sort_values(by=['day', 'num', 'date_pair'])

    # Группировка всех дат уникальных наборов "день - номер - предмет - препод - тип - группа - кабинет"
    dft = df.set_index(['day',
                        'num',
                        'item_name',
                        'teacher',
                        'type',
                        'pdgr',
                        'cab',
                        'date_pair']).groupby(level=[list(range(6))])

    # Формирование списка датафреймов уникальных наборов (для итерации по группам записей)
    dft = [rec.reset_index() for ind, rec in dft]

    days = iter(cr_def.DAYS_NAMES)  # Итератор по дням
    day = cr_def.DAYS_NAMES[0]      # Переход на первый день
    act_row = 1                  # Индекс актуальной строки на листе
    act_grp = 0                  # Индекс актуальной группы записей

    # Итерация по номерам пары для всех заданных дней
    for num in range(cr_def.TIMETABLE.shape[0] * len(cr_def.DAYS_NAMES)):
        # Актуальный номер пары в привычной записи
        act_num = num % cr_def.TIMETABLE.shape[0]

        # Новый день наступает после каждой итерации из семи пар (максимальное количество пар в день)
        if act_num < (num-1) % cr_def.TIMETABLE.shape[0]:
            day = next(days)

        # Итерироваться, если для текущего дня пар больше нет
        if act_grp == len(dft) or act_grp and dft[act_grp].loc[0, 'day'] != day:
            continue

        # Переход на новую строку
        act_row += 1

        # Занести на все листы текущий день, номер пары и её время
        # Стилизовать ячейки кабинетов и преподов
        for ws in wb:
            # Стилизовать и заполнить день недели
            _ = ws.cell(column=1, row=act_row, value=day).style = ST_DAYS

            # Стилизовать и заполнить номер пары
            _ = ws.cell(column=2, row=act_row, value=act_num + 1).style = ST_INFO

            # Время пары выбирается различно для выходных и рабочих дней
            if day != cr_def.DAYS_NAMES[5]:
                # Если будни
                _ = ws.cell(column=3, row=act_row, value=cr_def.TIMETABLE.loc[act_num + 1, 'weekdays']).style = ST_INFO
            else:
                # Если суббота
                _ = ws.cell(column=3, row=act_row, value=cr_def.TIMETABLE.loc[act_num + 1, 'weekends']).style = ST_INFO

            # Стилизовать ячейку преподов
            ws.cell(column=ws.max_column-1, row=act_row).style = ST_INFO

            # Стилизовать ячейку кабинетов
            ws.cell(column=ws.max_column,   row=act_row).style = ST_INFO

        # Проверить соответствие актуальной группы текущему дню и номеру пары
        if dft[act_grp].loc[0, 'day'] == day and dft[act_grp].loc[0, 'num'] == act_num+1:
            # Если соответствует, то обработать все записи для актуального дня и номера пары
            # Итерация по всем наборам записей для текущих дня и номера пары
            while True:
                # Итераторы для записи инфы о предмете
                weeks = tt.iterrows()      # Итератор по учебным неделям семестра
                act_week = next(weeks)[1]  # Актуал для проверяемой недели
                rec = None

                # Занесение информации в объект таблицы
                for ind, rec in dft[act_grp].iterrows():
                    # Если записи нет на этой учебной неделе: поиск нужной недели
                    while not act_week['monday'] <= rec['date_pair'] <= act_week['saturday']:
                        act_week = next(weeks)[1]

                    ws = wb[act_week['ind_sheet']]  # Актуальный лист
                    act_col = act_week['ind_col']   # Актуальный столбец на листе

                    # Дополнить ячейку инфой
                    # Форматная запись предмета
                    pat_rec = f"{rec['item_name']}: {rec['type']} ({rec['pdgr']})"
                    if ws.cell(column=act_col, row=act_row).value:
                        val = '\n'.join([ws.cell(column=act_col, row=act_row).value, pat_rec])
                    else:
                        val = pat_rec
                    # Стилизовать и заполнить ячейку инфы
                    _ = ws.cell(column=act_col, row=act_row, value=val).style = ST_COMMON

                    # Заполнение ячеек преподов и кабинетов
                    # Для корректного заполнения (умное добавление информации), ячейка разделяется
                    rzd_a = re.split(r'[\n,] ?', str(ws.cell(column=ws.max_column-1, row=act_row).value))
                    rzd_b = re.split(r'[\n,] ?', str(ws.cell(column=ws.max_column,   row=act_row).value))

                    # Добавление инфы о кабинете
                    # Случай 1) Пустая ячейка
                    if not ws.cell(column=ws.max_column-1, row=act_row).value:
                        val = rec['cab']
                        # Если кабинет - число, то привести его в числовой тип
                        if val.isdigit():
                            # Это нужно для отключения уведомлений EXCEL о числах, записанных строкой
                            val = int(val)
                        _ = ws.cell(column=ws.max_column-1, row=act_row, value=val)
                    # Случай 2) Препод не изменился, но кабинет другой
                    elif rzd_b[-1] == rec['teacher']:
                        if rzd_a[-1] != rec['cab']:
                            val = f"{ws.cell(column=ws.max_column - 1, row=act_row).value}, {rec['cab']}"
                            _ = ws.cell(column=ws.max_column-1, row=act_row, value=val)
                    # Случай 3) Кабинет изменился
                    else:
                        if rzd_a[-1] != rec['cab']:
                            val = f"{ws.cell(column=ws.max_column - 1, row=act_row).value}\n {rec['cab']}"
                            _ = ws.cell(column=ws.max_column-1, row=act_row, value=val)

                    # Добавление инфы о преподе
                    # Случай 1) Пустая ячейка
                    if not ws.cell(column=ws.max_column, row=act_row).value:
                        val = rec['teacher']
                        _ = ws.cell(column=ws.max_column, row=act_row, value=val)
                    # Случай 2) Препод изменился
                    elif not(rzd_b[-1] == rec['teacher']):
                        val = '\n'.join([ws.cell(column=ws.max_column, row=act_row).value, rec['teacher']])
                        _ = ws.cell(column=ws.max_column, row=act_row, value=val)

                # Перейти к следующему набору записей
                act_grp += 1

                # Если набор записей уже не соответствует текущим дню-номеру, то итерация завершается
                if rec is not None and act_grp == len(dft) or rec[0] != dft[act_grp].loc[0, 'day'] or rec[1] != dft[act_grp].loc[0, 'num']:
                    break

    # Вернуть книгу с заполненной инфой о всяком
    return wb


def is_merge(cell_1: 'Проверяемая ячейка №1',
             cell_2: 'Проверяемая ячейка №2',
             merges: 'Координатный список объединённых ячеек листа'
             ) -> 'Координаты для объединения ячеек, либо False':

    """ Функция умной проверки пересечения ячеек для объединения """

    # Во избежание попыток объединения пустых частей объединённых ячеек
    # Если обе ячейки пустые
    if not cell_1.value and not cell_2.value:
        # То их нельзя объединять
        return False

    # Определение типов ячеек (openpyxl не всегда корректно его предоставляет)
    t1 = t2 = None
    for cell_m in merges:
        if not t1 and cell_1.coordinate in cell_m:
            t1, cell_1 = 'MergedCell', re.findall(r'[A-Z](?=\d)', cell_m.coord)
        if not t2 and cell_2.coordinate in cell_m:
            t2, cell_2 = 'MergedCell', re.findall(r'[A-Z](?=\d)', cell_m.coord)

    # После определения, всегда есть четыре варианта ситуаций

    # Вариант 1) Обе ячейки - обычные. То есть, наложения между ними быть не может
    if not t1 and not t2:
        return [cell_1.coordinate[0], cell_2.coordinate[0]]

    # Группа случаев с объединёнными ячейками. Здесь нужно проверять равенство пограничных столбцов
    else:
        # Вариант 2) Обе ячейки - объединённые
        if t1 == t2:
            # Всё збс только если равенство столбцов
            if cell_1 == cell_2:
                return [cell_1[0], cell_1[1]]
            else:
                return False

        # Вариант 3) Первая ячейка - объединённая, вторая - обычная
        elif t1 and not t2:
            # Всё хорошо, если объединённая ячейка объединяет лишь одну ячейку
            if cell_1[0] == cell_1[1] == cell_2.coordinate[0]:
                return [cell_1[0], cell_1[1]]
            else:
                return False

        # Вариант 4) Первая ячейка - обычная, вторая - объединённая
        else:
            # Аналогично случаю три, но проверяется вторая ячейка
            if cell_2[0] == cell_2[1] == cell_1.coordinate[0]:
                return [cell_2[0], cell_2[1]]
            else:
                return False


def visual(wb:  'Частично форматированная книга'
           ) -> 'Полностью форматированная книга':

    """ Функция форматирования расписания """

    # Проход по каждому листу в книге
    for ws in wb:
        # Первый прогон - проверка "Есть ли пустые строки В КОНЦЕ дня"
        # Проверка идёт для всех строк, но без шапки и начиная с конца
        for row in range(ws.max_row, 1, -1):
            # Если нет кабинета и препода, а строка последняя, или у её нижнего соседа другой день:
            if all([not ws.cell(column=ws.max_column-1, row=row).value,
                    not ws.cell(column=ws.max_column,   row=row).value,
                    row == ws.max_row or ws.cell(column=1, row=row).value != ws.cell(column=1, row=row+1).value
                    ]):
                ws.delete_rows(row)  # Удалить строку

        # Второй прогон - выделение пустот
        # Прогон идёт по всем ячейкам с инфой о предметах
        for row in range(2, ws.max_row+1):
            for col in range(4, ws.max_column-1):
                # Если ячейка пустая, применить к ней стиль пустой ячейки
                if not ws.cell(column=col, row=row).value:
                    ws.cell(column=col, row=row).style = ST_NULL

        # Третий прогон - проверка "Слева - направо"
        for row in range(2, ws.max_row+1):
            # Пустая строка объединяется сразу
            if not ws.cell(column=4, row=row).value and not ws.cell(column=ws.max_column-1, row=row).value:
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ws.max_column-2)
                ws.cell(column=2, row=row).fill = ST_NULL.fill
                ws.cell(column=3, row=row).fill = ST_NULL.fill
                ws.cell(column=ws.max_column-1, row=row).fill = ST_NULL.fill
                ws.cell(column=ws.max_column,   row=row).fill = ST_NULL.fill
            else:
                sc = ec = 4
                for col in range(4, ws.max_column-1):
                    # Правая ячейка такая же как эта? Сдвинуть правую границу
                    if ws.cell(column=col, row=row).value == ws.cell(column=col+1, row=row).value:
                        ec += 1
                    # Нет? Вольтрон
                    else:
                        if sc != ec:
                            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
                        sc = ec = col+1

        # Четвёртый прогон - объединение дней и визуальное выделение дней
        # Начинается со столбца №2 из-за красивой рамки под временем и номером
        for col in range(2, ws.max_column+1):

            # Сброс нижних границ дня и объединения
            de = oe = ws.max_row
            sc = ec = 4

            # Проверка каждой строки, вплоть до шапки
            for row in range(ws.max_row, 1, -1):
                c1, c2 = ws.cell(column=col, row=row), ws.cell(column=col, row=row-1)
                coords = is_merge(c1, c2, ws.merged_cells.ranges)
                if coords:
                    sc, ec = ws[coords[0]][0].column, ws[coords[1]][0].column

                # Если день не сменился, то можно проверять то что внутри дня
                if ws.cell(column=1, row=row).value == ws.cell(column=1, row=row-1).value:
                    # Если верхняя ячейка не подходит для объединения, завершить выделение и начать новое
                    if c1.value != c2.value or not coords:
                        # Область выделения должна включать больше одной ячейки
                        if row != oe:
                            # openpyxl не объединяет уже объединённые ячейки (даже такие, которые можно объединить)
                            for rw in range(row, oe+1):
                                # Поэтому, всё в объединяемой области нужно разбить
                                try:
                                    ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                                except:
                                    pass
                            ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)
                        oe = row - 1

                # Если день сменился, то принудительно завершить выделение области + стилизовать день
                else:
                    # Объединить имеющуюся область выделения, если в ней нет проблем
                    if row != oe:
                        # openpyxl не объединяет уже объединённые ячейки (даже такие, которые можно объединить)
                        for rw in range(row, oe+1):
                            # Поэтому, всё в объединяемой области нужно разбить
                            try:
                                ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                            except:
                                pass
                        ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)

                    # Установить стилизованное выделение для границ дня
                    ws.cell(column=col, row=de).border = Border(left=ws.cell(column=col, row=de).border.left,
                                                                right=ws.cell(column=col, row=de).border.right,
                                                                top=ws.cell(column=col, row=de).border.top,
                                                                bottom=Side(border_style='thick'))

                    # Объединить день, но только когда обход уже на последнем столбце (иначе ненужное выделение)
                    if col == ws.max_column:
                        ws.merge_cells(start_row=row, start_column=1, end_row=de, end_column=1)

                    # Обновить нижние границы дня и объединения
                    de = oe = row-1

        # Пятый прогон - корректировка ширины столбца и высоты строки

        # Корректировка ширины столбцов
        for column in ws.columns:
            max_len = max(max((len(r) for r in re.split(r'\n', str(cell.value)))) for cell in column)
            max_len = (lambda m: m*2 if m < 10 else m*1.5 if m < 20 else m*1.45)(max_len)
            ws.column_dimensions[column[0].column_letter].width = max_len+1

        # Корректировка высоты строк
        for row in ws.rows:
            # Если строка пустая, высота минимальная
            if ws.cell(column=2, row=row[0].row).fill.start_color.index == ST_NULL_COLOR:
                ws.row_dimensions[row[0].row].height = 20
            # Если строка не пустая, высота вычисляемая
            else:
                max_str = max(len(re.split(r'\n', str(cell.value))) for cell in row) + 1
                ws.row_dimensions[row[0].row].height = max_str * 17

        # Закрепление столбцов слева и строк выше D2
        ws.freeze_panes = 'D2'

    # Вернуть отформатированную книгу
    return wb


def create_resp(df:   'База парсинга',
                grp2: 'Выбранная подгруппа (где две подгруппы)' = '0',
                grp3: 'Выбранная подгруппа (где три подгруппы)' = '0',
                i_yn: 'Сокращать запись предметов? (Аббревиатуры)' = False,
                t_yn: 'Сокращать запись преподов? (Без должности)' = False,
                p_yn: 'Сокращать запись подгрупп? (Без п/гр)' = False,
                c_yn: 'Сокращать запись кабинетов? (Без корпуса)' = False,
                ) -> 'Отформатированная книга расписания':

    """ Функция создания форматированного расписания из данных парсинга """

    """ Модификация базы данных """
    # Замена всех порченных типов пары на аналоги
    df = cr_add.replace_type(df, cr_add.find_bad_type(df))

    # Выборка базы данных для групп
    df = cr_add.take_data(df, grp2, grp3).dropna(subset=['date_pair'])

    # Выделение учебных недель
    times = time_resp(df)

    """ Форматирование базы данных """
    # Предметы
    if i_yn:
        df['item_name'] = df['item_name'].apply(lambda item_name: cr_add.format_item_name(item_name))

    # Преподы
    if t_yn:
        df['teacher'] = df['teacher'].apply(lambda teacher: cr_add.format_teacher(teacher))

    # Подгруппы
    if not p_yn:
        df['pdgr'] = df['pdgr'].apply(lambda pdgr: cr_add.format_pdgr(pdgr))

    # Кабинеты
    if c_yn:
        df['cab'] = df['cab'].apply(lambda cab: cr_add.format_cab(cab))

    """ Работа с EXCEL """
    # Создание новой таблицы
    wb = Workbook()

    # Заполнение типового шаблона таблицы
    wb = sheet_and_headers(wb, times)

    # Заполнение информационной части таблицы
    wb = fill_base(wb, df, times)

    # Форматирование информационной части таблицы
    wb = visual(wb)

    # Если всё ок, вернуть отформатированную таблицу
    return wb


def save_resp(book: 'Сохраняемая книга',
              path: 'Путь для сохранения'
              ) -> None:

    """ Функция сохраненения книги расписания в файл """

    book.save(path)
