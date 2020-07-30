""" Информация о модуле
Данный модуль предназначен для сохранения запарсенной базы расписания в виде отформатированной таблицы.
При сохранении учитываются результаты анализа БД, что позволяет минимизировать визуальные ошибки.

Примечание: openpyxl при обращении к ячейке сразу хранит её, из-за чего обработка ячейки удобна через _

Известные баги:
    1) Ширина столбцов некорректно рассчитывается для объединённых ячеек (т.к пока нет поправки на такие ячейки)

Что нужно не забыть добавить:
    1) Подсветка пар с наложением (а нужно ли это?)
    2) Удаление пустых учебных недель (на всякий случай)

"""

# Импорт команд для обработки дат из модуля datetime
from datetime import date as dt_date, timedelta

# Импорт базовых команд поиска и замены из стандартного модуля регулярных выражений
from re import findall, fullmatch, match, search, split, sub

# Импорт команд обработки ячеек EXCEL из модуля openpyxl
from openpyxl import Workbook

# Импорт команд форматирования ячеек EXCEL из модуля openpyxl
from openpyxl.styles import NamedStyle, Alignment, Border, Font, GradientFill, PatternFill, Side


""" Константные стили и названия """
# Стиль шапки
st_title = NamedStyle(name = 'Шапка')
st_title.font = Font(name = 'Book Antiqua', size = 14)
st_title.alignment = Alignment(horizontal = 'center', vertical = 'center')
st_title.border = Border(left = Side(border_style = 'medium'),  right = Side(border_style = 'medium'),
                         top  = Side(border_style = 'medium'), bottom = Side(border_style = 'medium'))

# Стиль дней
st_days = NamedStyle(name = 'Дни')
st_days.font = Font(name = 'Bookman Old Style', size = 14, bold = True)
st_days.alignment = Alignment(horizontal = 'center', vertical = 'center', textRotation = 90)
st_days.border = Border(left = Side(border_style = 'medium'), right  = Side(border_style = 'medium'),
                        top  = Side(border_style = 'medium'), bottom = Side(border_style = 'medium'))

# Стиль базовой ячейки
st_baze = NamedStyle(name = 'Базовая ячейка')
st_baze.font = Font(name = 'Plantagenet Cherokee', size = 14)
st_baze.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
st_baze.border = Border(left = Side(border_style = 'thin'), right  = Side(border_style = 'thin'),
                        top  = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))

# Стиль для номеров пары и времени
st_razn = NamedStyle(name = 'Инфополе')
st_razn.font = Font(name = 'Plantagenet Cherokee', size = 14)
st_razn.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
st_razn.border = Border(left = Side(border_style = 'medium'), right  = Side(border_style = 'medium'),
                        top  = Side(border_style = 'thin'),   bottom = Side(border_style = 'thin'))

# Стиль пустой ячейки
st_null = NamedStyle(name = 'Круговерть пустоты')
st_null.border = Border(left = Side(border_style = None), right  = Side(border_style = None),
                        top  = Side(border_style = None), bottom = Side(border_style = None))
st_null.fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')

# Названия столбцов
parse_title = ['Дни', '№ пары', 'Время', 'Ауд', 'Преподаватель']

# Полные названия месяцев
mos = {1: 'Январь',    2: 'Февраль',  3: 'Март',    4: 'Апрель',
       5: 'Май',       6: 'Июнь',     7: 'Июль',    8: 'Август',
       9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}

# Сокращения месяцев (альтернатива "mos[date[0].month][0].lower()+mos[date[0].month][1:3]")
# Улучшает читаемость кода + повышает кастомизацию
mos_s = {1: 'янв',   2: 'фев',  3: 'марта', 4: 'апр',
         5: 'мая',   6: 'июня', 7: 'июля',  8: 'авг',
         9: 'сент', 10: 'окт', 11: 'нояб', 12: 'дек'}

# Расписание звонков в будние дни. 5-минутные перерывы не учитываются, т.к преподы их не делают
time_budni = {0: '08:30 - 10:05', 1: '10:15 - 11:50', 2: '12:35 - 14:10',
              3: '14:20 - 15:55', 4: '16:05 - 17:40', 5: '17:50 - 19:25'}

# Расписание звонков в субботу. 5-минутные перерывы не учитываются, т.к преподы их не делают
time_vihod = {0: '08:30 - 10:05', 1: '10:15 - 11:50', 2: '12:00 - 13:40',
              3: '13:50 - 15:25', 4: '15:35 - 17:10', 5: '17:20 - 18:55'}

""" Функция определения периода для дней в шапке
Описание:
    Берёт период расписания и преобразует его в список кортежей "начало недели - конец недели"

Аргументы:
    1) Период расписания
    2) Год расписания

Возвращаемое:
    Список из объектов datetime - границ (с ПН по СБ) учебных недель

"""
def time_resp(period, year):
    # Список для хранения границ всех учебных недель (элемент - кортеж с границами одной недели)
    tdt = []

    dt_start = list(map(int, period[0].split('.')))
    dt_final = list(map(int, period[1].split('.')))
    dt_start = dt_date(year, dt_start[1], dt_start[0])
    dt_final = dt_date(year, dt_final[1], dt_final[0])

    s = 0 # Начало учебной недели
    while dt_start <= dt_final:
        # Если добавляется первая или новая неделя
        if s == 0 or dt_start.weekday() == 0:
            s = dt_start
        # Если уже суббота, нужно пропустить воскресенье
        elif dt_start.weekday() == 5:
            tdt.append([s, dt_start])
            dt_start += timedelta(days = 1)
        dt_start += timedelta(days = 1)

    # Возврат списка границ учебных недель
    return tdt

""" Функция заполнения листов таблицы и их шапок
Описание:
    Функция, для каждого учебного месяца, создаёт свой лист, на котором заполняются:
    1) Шапка (названия столбцов)
    2) Временные рамки

Аргументы:
    1) Объект книги EXCEL из openpyxl
    2) Список границ учебных недель

Возвращаемое:
    Изменённый объект книги из openpyxl

"""
def sheet_and_headers(wb, tdt):
    # Заполнение листов таблицы и их шапок
    for d_ind, date in enumerate(tdt):
        # Учебный месяц изменяется при определённом наборе условий
        if any([all([d_ind,
                     any([tdt[d_ind-1][0].month != date[0].month,
                          tdt[d_ind-1][1].month != date[1].month]),
                     any([int(date[0].day) in range(1, 7),
                          int(date[1].day) in range(1, 7)]),
                     mos[date[0].month] not in wb.sheetnames]),
                not d_ind]):

            # Если в месяце была одна неделя (при смене месяца), то нужно просто объединить месяцы
            if not wb.sheetnames[0] == 'Sheet' and ws.max_column < 5:
                ws.title = mos[date[0].month]
            # Если в следующем месяце будет всего одна неделя (на случай чумы, апокалипсиса или сессии), то ничего не менять
            elif not(d_ind+1 == len(tdt) or d_ind and tdt[d_ind-1][0].month != date[0].month and date[0].month != tdt[d_ind+1][0].month):
                # Если добавляется первый лист, то можно просто переименовать стандартный
                if wb.sheetnames[0] == 'Sheet':
                    ws = wb.active
                    ws.title = mos[date[0].month]
                # Если обычный новый учебный месяц, то создать новый лист
                else:
                    wb.create_sheet(mos[date[0].month])

                # Сделать актуальным листом последний лист
                ws = wb.worksheets[-1]
                # Заполнение первых трёх столбцов шапки
                _ = ws.cell(column=ws.max_column,   row=1, value=parse_title[0]).style = st_title
                _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[1]).style = st_title
                _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[2]).style = st_title

        # Запись учебной недели в красивом формате
        m1, m2 = mos_s[date[0].month], mos_s[date[1].month]
        val = date[0].strftime('%d') + ' ' + m1 + ' - ' + date[1].strftime('%d') + ' ' + m2
        _ = ws.cell(column=ws.max_column+1, row=1, value=val).style = st_title

        # Если последняя учебная неделя, или произошла смена листа, то доставить титулы кабинетов и преподов
        if d_ind == len(tdt)-1 or ws.max_column == 4 and len(wb.worksheets) > 1:
            if d_ind == len(tdt)-1:
                pred = wb.worksheets[-1]
            else:
                pred = wb.worksheets[-2]
            _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[3]).style = st_title
            _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[4]).style = st_title

    # Вернуть книгу с готовым шаблоном
    return wb
  

""" Функция заполнение основной части таблицы
Описание:
    Здесь происходит заполнение таблицы инфой из запарсенной базы

Аргументы:
    1) Объект книги EXCEL из openpyxl
    2) База парсинга
    3) База анализа
    4) Список границ учебных недель
    5) Подгруппы, для которых составляется расписание
    6) Список умолчаний для дней

Возвращаемое:
    Книга с инфой из базы парсинга

"""
def fill_base(wb, parse, a_bd, tdt, grps, days_names):
    # Константные установки и основные индексы
    grps = {1: 0, 2: grps[0], 3: grps[1]} # Для выбора подгруппы

    days = iter(days_names) # Итератор по дням
    act_r = 1  # Индекс актуальной строки в таблицы
    act_bd = 0 # Актуальная запись в базе разбора

    # Прогон по номерам пары в шаблоне, чтобы день не мог получиться таким как "№2, №4, №6"
    for num in range(len(time_budni) * 6):
        # Если следующий день
        if num%6 < (num-1)%6:
            day = days_names[next(days)]
        # Если для дня пары закончились, но время ещё есть - лучше просто перейти к следующему дню
        if act_bd == len(parse) or act_bd and parse[act_bd][0] != day:
            continue
        act_r += 1

        # Занести день, номер и время пары на все листы. Применить "разный" стиль к кабинетам и преподам, мало ли где пропуск
        for ws in wb:
            _ = ws.cell(column = 1, row = act_r, value = day).style = st_days
            _ = ws.cell(column = 2, row = act_r, value = num%6 + 1).style = st_razn
            if day != days_names[5]:
                _ = ws.cell(column = 3, row = act_r, value = time_budni[num % 6]).style = st_razn
            else:
                _ = ws.cell(column = 3, row = act_r, value = time_vihod[num % 6]).style = st_razn
            ws.cell(column = ws.max_column-1, row = act_r).style = st_razn
            ws.cell(column = ws.max_column,   row = act_r).style = st_razn

        # Проверить соответствие записи в БД актуальному дню и номеру пары
        if not act_bd or parse[act_bd][0] == day and parse[act_bd][1] == num%6+1:
            # Если соответствует, то обработать все записи для актуального дня и номера пары
            while True:
                # Итераторы для записи инфы о предмете
                weeks = iter(tdt)               # Итератор по учебным неделям семестра
                it_week = next(weeks)           # Актуал для проверяемой недели
                it_sheets = iter(wb.sheetnames) # Итератор по листам книги
                ws = wb[next(it_sheets)]        # Актуальный лист
                act_c = 4                       # Актуальный столбец. Для проверки на смену листа
                rec = parse[act_bd]             # Актуальная запись в базе разбора

                # Проверить подгруппу у типа пары в графе замены
                zam = a_bd[rec[2]][rec[4]][1]   # На что заменить тип пары
                gr = grps[a_bd[rec[2]][zam][0]] # Выбранная подгруппа для типа пары

                # Если на подгруппу пофиг (0), или подгруппа соответствует текущей, то инициировать занесение информации в таблицу
                if not gr or gr == int(rec[5][0]):
                    # Прогон по каждой дате в записи
                    for date in rec[6]:
                        # Если дата не входит в период расписания
                        if date < tdt[0][0] or date > tdt[-1][1]:
                            continue
                        # Если записи нет на этой учебной неделе: поиск нужной недели, т.е...
                        while not (it_week[0] <= date <= it_week[1]):
                            it_week = next(weeks) #...переход к новой неделе
                            act_c += 1  #...переход на следующий столбец таблицы
                            # Проверить, есть ли новая учебная неделя на текущем листе: если актуал столбца на кабинетном столбце, то...
                            if act_c > ws.max_column-2:
                                ws = wb[next(it_sheets)] #...переход на следующий лист
                                act_c = 4                #...обновление актуального столбца

                        # Добавление инфы о предмете и подгруппе
                        pat_rec = f'{rec[2]}: {rec[4]} ({rec[5]})' # Форматная запись предмета
                        if ws.cell(column=act_c, row=act_r).value:
                            val = '\n'.join([ws.cell(column=act_c, row=act_r).value, pat_rec])
                        else:
                            val = pat_rec
                        _ = ws.cell(column=act_c, row=act_r, value=val).style = st_baze

                        rzd_a = split(r'[\n,] ?', str(ws.cell(column=ws.max_column-1, row=act_r).value))
                        rzd_b = split(r'[\n,] ?', str(ws.cell(column=ws.max_column,   row=act_r).value))

                        if not ws.cell(column=ws.max_column, row=act_r).value or rzd_b[-1]!=rec[3] or rzd_b[-1]==rec[3] and rzd_a[-1]!=rec[7]:
                            # Добавление инфы о кабинете
                            # Случай 1) Пустая ячейка
                            if not ws.cell(column=ws.max_column-1, row=act_r).value:
                                val = rec[7]
                            # Случай 2) Препод не изменился, но кабинет другой
                            elif rzd_b[-1] == rec[3]:
                                val = ', '.join([ws.cell(column=ws.max_column-1, row=act_r).value, rec[7]])
                            # Случай 3) Не пустая ячейка
                            else:
                                val = '\n'.join([ws.cell(column=ws.max_column-1, row=act_r).value, rec[7]])
                            _ = ws.cell(column=ws.max_column-1, row=act_r, value=val)

                            # Добавление инфы о преподе
                            # Случай 1) Пустая ячейка
                            if not ws.cell(column=ws.max_column, row=act_r).value:
                                val = rec[3]
                            # Случай 2) Препод не изменился, но кабинет другой
                            elif not(rzd_b[-1] == rec[3]):
                                val = '\n'.join([ws.cell(column=ws.max_column, row=act_r).value, rec[3]])
                            else:
                                continue
                            _ = ws.cell(column=ws.max_column, row=act_r, value=val)

                act_bd += 1
                if act_bd == len(parse) or rec[0]!=parse[act_bd][0] or rec[1]!=parse[act_bd][1]:
                    break

        # Форматирование инфы о кабинете (форматируется после заполнения инфы)
        for ws in wb:
            cab = ws.cell(column = ws.max_column-1, row = act_r).value
            if cab:
                cab = set(cb for cb in split(r'[\n,] ?', cab) if cb)
                # Если инфа о кабинетах - один дублирующийся кабинет, то схлопнуть её
                if len(cab) == 1:
                    cab = list(cab)[0]
                    try: # Эксель порой ругается, что "число сохранено как текст". Бесит
                        cab = int(cab)
                    except:
                        pass
                    ws.cell(column = ws.max_column-1, row = act_r).value = cab

    # Вернуть книгу с заполненной инфой о всяком
    return wb


""" Функция для проверки ячеек на пересечение
Описание:
    Ячейки нельзя объединять не в форме прямоугольной области выделения
    То есть, перед объединением всё нужно проверять

Аргументы:
    1) Проверяемая ячейка №1
    2) Проверяемая ячейка №2
    3) Координатный список объединённых ячеек листа

Возвращаемое:
    1) Если ячейки можно объединить, то координатный список для объединения ячейки (столбец 1, столбец 2)
    2) Если нет, то False

"""
def cross_merge(cell_1, cell_2, merge_list):
    # Во избежание попыток объединения пустых частей объединённых ячеек
    if not cell_1.value and not cell_2.value:
        return False
    # Определение типов ячеек (модуль не всегда корректно его предоставляет, хз почему, но доказано логами и крашами)
    t1 = t2 = None
    for cell_m in merge_list:
        if not t1 and cell_1.coordinate in cell_m:
            t1, cell_1 = 'MergedCell', findall(r'[A-Z](?=\d)', cell_m.coord)
        if not t2 and cell_2.coordinate in cell_m:
            t2, cell_2 = 'MergedCell', findall(r'[A-Z](?=\d)', cell_m.coord)
    # После определения, всегда есть четыре варианта ситуаций
    # Вариант 1) Обе ячейки - обычные. То есть, наложения между ними быть не может
    if not t1 and not t2:
        return [cell_1.coordinate[0], cell_2.coordinate[0]]
    # Группа случаев с объединёнными ячейками. Здесь нужно проверять равенство пограничных столбцов
    else:
        # Вариант 2) Обе ячейки - объединённые
        if t1 == t2:
            # Всё збс только если равенство столбцов
            if cell_1 == cell_2:
                return [cell_1[0], cell_1[1]]
            else:
                return False
        # Вариант 3) Первая ячейка - объединённая, вторая - обычная
        elif t1 and not t2:
            # Всё збс если каким-то фиговым листом объединённая ячейка объединяет всего одну ячейку
            if cell_1[0] == cell_1[1] == cell_2.coordinate[0]:
                return [cell_1[0], cell_1[1]]
            else:
                return False
        # Вариант 4) Первая ячейка - обычная, вторая - объединённая
        else:
            # Аналогично случаю три, но проверяется вторая ячейка
            if cell_2[0] == cell_2[1] == cell_1.coordinate[0]:
                return [cell_2[0], cell_2[1]]
            else:
                return False


""" Функция донастройки визуала таблицы
Описание:
    Конечное форматирование таблицы

Примечание:
    Форматирование идёт в несколько прогонов, чтобы не поймать пересечение форматирования

Аргументы:
    Частично форматированная книга

Возвращаемое:
    Полностью форматированная книга

"""
def visual(wb):
    # Проход по каждому листу в книге
    for ws in wb:
        # Первый прогон - проверка "Есть ли пустые строки В КОНЦЕ дня"
        # Проверка идёт для всех строк, но без шапки и начиная с конца
        for row in range(ws.max_row, 1, -1):
            # Если нет кабинета и препода, а строка последняя, или у её нижнего соседа другой день:
            if all([not ws.cell(column=ws.max_column-1, row=row).value,
                    not ws.cell(column=ws.max_column  , row=row).value,
                    row == ws.max_row or ws.cell(column=1, row=row).value!=ws.cell(column=1, row=row+1).value
                   ]):
                ws.delete_rows(row)# Удалить строку

        # Второй прогон - выделение пустот
        # Прогон идёт по всем ячейкам с инфой о предметах
        for row in range(2, ws.max_row+1):
            for col in range(4, ws.max_column-1):
                # Если ячейка пустая, применить к ней стиль пустой ячейки
                if not ws.cell(column=col, row=row).value:
                    ws.cell(column=col, row=row).style = st_null

        # Третий прогон - проверка "Слева - направо"
        for row in range(2, ws.max_row+1):
            # Пустая строка объединяется сразу
            if not ws.cell(column=4, row=row).value and not ws.cell(column=ws.max_column-1, row=row).value:
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ws.max_column-2)
                ws.cell(column = 2, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
                ws.cell(column = 3, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
                ws.cell(column = ws.max_column-1, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
                ws.cell(column = ws.max_column,   row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
            else:
                sc = ec = 4
                for col in range(4, ws.max_column-1):
                    # Правая ячейка такая же как эта? Сдвинуть правую границу
                    if ws.cell(column=col, row=row).value == ws.cell(column=col+1, row=row).value:
                        ec += 1
                    # Нет? Вольтрон
                    else:
                        if sc != ec:
                            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
                        sc = ec = col+1

        # Четвёртый прогон - объединение дней и визуальное выделение дней. Начинается со столбца №2 из-за красивой рамки под временем и номером
        for col in range(2, ws.max_column+1):

            # Сброс нижних границ дня и объединения
            de = oe = ws.max_row
            sc = ec = 4

            # Проверка каждой строки, вплоть до шапки
            for row in range(ws.max_row, 1, -1):
                c1, c2 = ws.cell(column=col, row=row), ws.cell(column=col, row=row-1)
                coords = cross_merge(c1, c2, ws.merged_cells.ranges)
                if coords:
                    sc, ec = ws[coords[0]][0].column, ws[coords[1]][0].column

                # Если день не сменился, то можно проверять то что внутри дня
                if ws.cell(column=1, row=row).value == ws.cell(column=1, row=row-1).value:
                    # Если верхняя ячейка не подходит для объединения, завершить выделение и начать новое
                    if c1.value != c2.value or not coords:
                        if row != oe: # Область выделения должна включать больше одной ячейки
                            # Openpyxl немного туповат и не объединяет объединённые ячейки (даже такие, которые можно объединить)
                            for rw in range(row, oe+1):
                                # Поэтому, всё в объединяемой области нужно разбить
                                try:
                                    ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                                except:
                                    pass
                            ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)
                        oe = row - 1

                # Если день сменился, то принудительно завершить выделение области + стилизовать день
                else:

                    # Объединить имеющуюся область выделения, если в ней нет проблем
                    if row != oe:
                        # Openpyxl немного туповат и не объединяет объединённые ячейки (даже такие, которые можно объединить)
                        for rw in range(row, oe+1):
                            # Поэтому, всё в объединяемой области нужно разбить
                            try:
                                ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                            except:
                                pass
                        ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)

                    # Установить стилизованное выделение для границ дня
                    ws.cell(column=col, row=de).border = Border(left   = ws.cell(column=col, row=de).border.left,
                                                                right  = ws.cell(column=col, row=de).border.right,
                                                                top    = ws.cell(column=col, row=de).border.top,
                                                                bottom = Side(border_style = 'medium'))

                    # Объединить день, но только когда обход уже на последнем столбце (иначе творится дичь + ненужное выделение)
                    if col == ws.max_column:
                        ws.merge_cells(start_row=row, start_column=1, end_row=de, end_column=1)

                    # Обновить нижние границы дня и объединения
                    de = oe = row-1

        # Пятый прогон - корректировка ширины столбца и высоты строки
        for column in ws.columns: # Ширина столбца
            max_len = max(max((len(r) for r in split(r'\n', str(cell.value)))) for cell in column)
            max_len = (lambda m: m*2 if m<10 else m*1.5 if m<20 else m*1.45)(max_len)
            ws.column_dimensions[column[0].column_letter].width = max_len+1

        for row in ws.rows:       # Высота строки
            max_str = max(len(split(r'\n', str(cell.value))) for cell in row) + 1
            ws.row_dimensions[row[0].row].height = max_str * 17
        ws.freeze_panes = 'D2'

    # Вернуть отформатированную книгу
    return wb


""" Функция создания форматированного расписания
Описание:
    Создаёт и сохраняет форматированное расписание

Аргументы:
    1) База парсинга
    2) База анализа
    3) Подгруппы, для которых составляется расписание
    4) Путь к месту сохранения
    5) Умолчания для дней
    6) Период расписания
    7) Год расписания

Возвращаемое:
    -

"""
def create_resp(parse, a_bd, grps, name, days, interval, year):
    # Создание новой таблицы
    wb = Workbook()

    # Выделение учебных недель
    times = time_resp(interval, year)

    # Заполнение типового шаблона таблицы
    wb = sheet_and_headers(wb, times)

    # Заполнение информационной части таблицы
    wb = fill_base(wb, parse, a_bd, times, grps, days)

    # Форматирование информационной части таблицы
    wb = visual(wb)

    # Сохранение таблицы по указанному пути
    wb.save(name)
