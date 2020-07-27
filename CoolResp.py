# Коробка с костылями, выкованная дворфами из грязнейшего говнокода и зачарованная эльфийскими регулярными выражениями
# Для считывания таблицы используется pyexcel, для записи и форматирования - openpyxl
# Почему так? Первый модуль больше подходит для выдёргивания данных, а второй - для аккуратной их записи
# + pyexcel считывает как .xls, так и .xlsx (а также веб-ресурсы), а openpyxl поддерживает только .xlsx

# Импорт команды для считывания EXCEL таблицы из модуля pyexcel
from pyexcel import get_book_dict as pxl_book

# Импорт модуля регулярных выражений. Используется большинство его команд, т.е лучше импортировать полностью
import re

import requests

# Импорт команд для обработки дат из модуля datetime
from datetime import date as dt_date, timedelta as dt_tdelta

# Импорт команд обработки ячеек EXCEL из модуля openpyxl
from openpyxl import Workbook

# Импорт команд форматирования ячеек EXCEL из модуля openpyxl
from openpyxl.styles import NamedStyle, Alignment, Border, Font, GradientFill, PatternFill, Protection, Side

""" Считывание файла в книжный словарь """
# Тесты с несколькими группами и листами
#temp_table = pxl_book(file_name='Тесты/1_KURS_2018-2019_2semestr.xls')     # 1 Интересный тест
#temp_table = pxl_book(file_name='Тесты/2_kurs_2semestr_2018-2019.xls')     # 2  
#temp_table = pxl_book(file_name='Тесты/3_kurs_2018-2019_II_semestr.xls')   # 3  
#temp_table = pxl_book(file_name='Тесты/3_kurs_2019-2020_II_semestr.xls')   # 4  
#temp_table = pxl_book(file_name='Тесты/4_kurs_2018-2019_II_semestr.xls')   # 5  
#temp_table = pxl_book(file_name='Тесты/PE-91_92b_2_semestr_2019-2020.xls') # 6  

# Тесты где всего один лист и всего одна группа
#temp_table = pxl_book(file_name='Тесты/921_2_semestr_2019-2020.xls')       # 7  Интересный тест 
#temp_table = pxl_book(file_name='Тесты/MITE-91_2019-2020_II_semestr.xls')  # 8  
#temp_table = pxl_book(file_name='Тесты/MIVT-91_2019-2020_II_semestr.xls')  # 9  
#temp_table = pxl_book(file_name='Тесты/PE-51b_2018-2019_II_semestr.xlsx')  # 10 
#temp_table = pxl_book(file_name='Тесты/PE-61b_2018-2019_II_semestr.xlsx')  # 11 Интересный тест 
#temp_table = pxl_book(file_name='Тесты/PE-71b_2019-2020_II_semestr.xlsx')  # 12 
#temp_table = pxl_book(file_name='Тесты/PE-71_2_semestr_2018-2019.xls')     # 13 Белкина
#temp_table = pxl_book(file_name='Тесты/PE-81_2_semestr_2018-2019.xls')     # 14 
temp_table = pxl_book(file_name='Тесты/PE-81_2_semestr_2019-2020.xls')     # 15 Интересный тест

# Тесты веба
#h_table = requests.get('https://vk.com/doc33496599_544927498?hash=e200ba5c080bedb616&dl=7776342d094c9f496f')
#temp_table = pxl_book(file_content=h_table.content)

""" Выбор листа с книги """
if len(temp_table) > 1:
    # Если листов несколько, то нужно выбрать один из них
    #print(*temp_table.keys(), sep=' | ')
    #sheet = input('Выберите лист: ')
    sheet = 'МЕ,ОЕ 71б' # Для 2
    #sheet = 'ИТ,ВЕ-71б'
    #sheet = 'МЕ,ОЕ-61б' # Для 3
    #sheet = 'ИТ,ВЕ-61б'
    #sheet = 'МЕ,ОЕ-71б' # Для 4
    #sheet = 'ИТ,ВЕ-71б'
    #sheet = 'МЕ,ОЕ-51б' # Для 5
    #sheet = 'ИТЕ,ВЕ-51б'
else:
    # Если лист один, то автовыбор его
    for i in temp_table:
        sheet = i

""" Выделение позиции графика расписания на табличном листе
Принцип таков:
1) Корректное расписание всегда начинается через одну строку после указания его периода
2) Корректное расписание всегда заканчивается упоминанием должности "Начальник УО"
3) Т.е, расписание - всё что внутри 1-2, но начиная с 3-й строки
"""
for i, a in enumerate(temp_table[sheet]):
    if re.search(' *[Нн]а *период', a[0], re.A):
        start = i
    if re.search(' *[Нн]ачальник *УО', a[0], re.A):
        end = i

""" Загрузить график расписания групп на листе, и выгрузить книгу EXCEL из памяти """
temp_sheet = temp_table[sheet][start:end]
del(temp_table)

""" Удаление пустых столбцов (а они иногда встречаются) """
if '' in temp_sheet[0]:
    temp_sheet = list(zip(*temp_sheet))
    temp_sheet = [rs for rs in temp_sheet if any(rs)]
    temp_sheet = [list(row) for row in zip(*temp_sheet)]

""" Обособление периода расписания и групп """
# Период выделяется из "с <дата1> по <дата2>" в кортеж (дата1, дата2)
timey_wimey = re.findall(r'(с\s*[\d.]{5,8}[г.]*\s*по\s*[\d.]{5,8}[г.]*)', temp_sheet[0][0])[0]
period = re.findall(r'с\s*([\d.]{5,8})[г.]*\s*по\s*([\d.]{5,8})[г.]*', temp_sheet[0][0])[0]
year = str(2000+int(period[0][-2:]))

# Строка групп сохраняется как список групп на листе
#title = [temp_sheet[1][i] for i in range(2, len(temp_sheet[1]), 2)]
#print(f'Список групп на листе: {title}')
title = temp_sheet[1] # Строка групп

# Из расписания удаляются строки периода и списка групп
temp_sheet = temp_sheet[2:]


""" Процедура корректного считывания объединённых ячеек """
def merged_cells(row, col, cabs):
    act = col-2
    # Если есть предмет и правый сосед не кабинет - ячейка общая
    if row[act] and not row[act+1]:
        return [row[0], row[1], row[act], cabs]
    # Если ячейка 100% не общая или строка пройдена - вернуть пустую запись
    elif row[act+1] or act == 2 and not row[act]:
        return[row[0], '', '', '']
    else:
        return merged_cells(row, act, cabs)

""" Процедура отсеивания ошибок в ячейках (начальная стадия разбора) """
# Сокращённые записи дней. Глобальная переменная, так как используется при сохранении
days_names = {0: 'ПН', 1: 'ВТ', 2: 'СР', 3: 'ЧТ', 4: 'ПТ', 5: 'СБ'}
def errors_clear(sheet, i, row, act=-2):
    days_pat = r'(.*[лЛ].*)|(.*[вВ].*[оО].*[кК].*)|(.*[сС].*[дД].*)|(.*[чЧ].*[гГ].*)|(.*[яЯ].*)|(.*[уУ].*[бБ].*)'   
    if not row[0]: # Т.к день - общая ячейка
        row[0] = sheet[i-1][0]
    else: # Если день есть, привести к общему виду
        for i, sovp in enumerate(re.findall(days_pat, row[0])[0]):
            if sovp:
                row[0] = days_names[i]
                break
    if not row[1]: # Один номер пары порой включает несколько строк
        row[1] = sheet[i-1][1]
    """ Чистка кабинетов и приведение их к общему виду """
    # Начальные замены "до разделения"
    def repl_a(cab):
        # Единый формат для актового зала
        cab = re.sub(r'[аА].*?[лЛ]', 'АКТ_зал', cab)
        # Каб. для ФЗК в разных расписаниях может иметь вид: "с/з, т/з", "1xx УК№1 с/з, т/з", "1xx УК№1"
        pat_zal = r'(?:с\s*?/\s*?з.*?т\s*?/\s*?з)|(?:1[\d]{2}\s*?[уУ][кК]\s*?№\s*?1[,;: ].*т\s*?/\s*?з)|(?:1[\d]{2}\s*?[уУ][кК]\s*?№\s*?1)'
        cab = re.sub(pat_zal, 'ФЗК_зал,', cab)
        cab = re.sub(r'с.+?об', '', cab)
        cab = re.sub(r'\n+', ' ; ', cab)
        cab = re.sub(r'\s+', ' ', cab)
        return cab
    # Замены при формировании списка кабинетов
    def repl_b(cab):
        cab = re.sub(r'[.;: ]', '', cab)
        cab = re.sub(r'[уУ][кК]?№', ' УК№', cab)
        return cab
    if row[act+1]:
        cabs = repl_a(row[act+1])
        cabs = [repl_b(cb) for cb in re.split(r'[;,]', cabs)
                if not re.fullmatch(r'[.,:; ]*', cb)]
        # Исправление ['210', '212 УК№1', '329', '331 УК№5', '410 УК№1']
        for i_cb in range(len(cabs)-1, -1, -1):
            if re.match(r'\d+', cabs[i_cb]) and not re.match(r'\d+ УК№\d', cabs[i_cb]):
                if re.search(r' УК№\d', cabs[i_cb+1]):
                    cabs[i_cb] += re.search(r' УК№\d', cabs[i_cb+1])[0]
                else:
                    cabs[i_cb] += ' УК№?'
        row[act+1] = cabs
    return row

""" Процедура замены зачётов/дифов (мешают обработке) """
def swap_quiz(inf_data):
    pattern_dif = r'(?:зач[её]т[\s]*с[\s]*оценкой)|(?:диф[.\s]*зач[её]т)' # Отлов диф.зачётов
    pattern_zac = r'зач[её]т' # Отлов зачётов
    inf_data = re.sub(r'\s+', ' ', inf_data) # Пробельная чистка
    if re.search(pattern_dif, inf_data, flags = re.I): # Замена дифов на 6D6D6D6
        inf_data = re.sub(pattern_dif, '6D6D6D6', inf_data, flags = re.I)
    if re.search(pattern_zac, table[-1][2], flags = re.I): # Замена зачётов на 7Z7Z7Z7
        inf_data = re.sub(pattern_zac, '7Z7Z7Z7', inf_data, flags = re.I)
    return inf_data


""" Создание базы обработки (table) для одной из групп на листе
По сути, это контейнер с мусором, из которого хлам отправится на дальнейшую переработку
* Структура элемента формируемой базы обработки следующая:
    [0] - день
    [1] - номер пары
    [2] - ячейка с инфой по предметам, преподам, группам, и т.п
    [3] - список кабинетов
"""
# Для наглядности доп. условий, база обработки создаётся циклом а не через List Comprehension
table = []
# Таблицы сравнительно малы => можно хранить пустые строки (нужны для корректных дней, № пар)
if len(title) > 4: # Если группа одна, титул состоит из 4-х ячеек. Иначе - 4+
    # Выделение списка групп на листе
    g = {title[a]:a for a in range(2, len(title), 2)}
    #print(*g.keys(), sep=' | ')
    #v = g[input('Выберите группу: ')]
    #v = g['МЕ-81б'] # Для 1
    #v = g['ИТ-81б']
    #v = g['ОЕ-81б']
    #v = g['МЕ-71б'] # Для 2
    #v = g['ОЕ-71б']
    #v = g['ИТ-71б']
    #v = g['ВЕ-71б']
    #v = g['МЕ-61б'] # Для 3
    #v = g['ОЕ-61б']
    #v = g['ИТ-61б']
    #v = g['ВЕ-61б']
    #v = g['МЕ-71б'] # Для 4
    #v = g['ОЕ-71б']
    #v = g['ИТ-71б']
    #v = g['ВЕ-71б']
    #v = g['МЕ-51б'] # Для 5
    #v = g['ОЕ-51б']
    v = g['ИТ-51б']
    #v = g['ВЕ-51б']
    #v = g['ПЕ-91б'] # Для 6
    #v = g['ПЕ-92б']
    # Каждая строка расписания добавляется на разбор в формате [День, №, инф_ячейка, кабинеты]
    for ind, record in enumerate(temp_sheet):
        record = errors_clear(temp_sheet, ind, record, v)
        # Когда групп несколько, встречаются потоковые лекции (общие ячейки)
        if record[v] or record[v-1]:
            table.append([record[0], record[1], record[v], record[v+1]])
        else:
            table.append(merged_cells(record, v, record[v+1]))
        if table[-1][2]: # Стандарт пробельных символов в информации
            table[-1][2] = swap_quiz(table[-1][2])
else: # Если группа всего одна
    v = 2 # Столбец предмета в пайтоновской нумерации
    for ind, record in enumerate(temp_sheet):
        record = errors_clear(temp_sheet, ind, record)
        # В расписании одной группы нет общих ячеек с информацией о предмете
        table.append([record[0], record[1], record[v], record[v+1]])
        if table[-1][2]: # Стандарт пробельных символов в информации
            table[-1][2] = swap_quiz(table[-1][2])

""" Умолчания """
default_grup = 'общ'  # Обозначение для общей пары
tip_list = {0: 'ДИФ', 1: 'зачёт', 2: 'лекция', 3: 'лекция',  4: 'ЛБ', 5: 'ПР'}
default_type_a = 'рейд'  # Обозначение для предмета без типа пары и подгруппы
default_type_b = 'ЛБ/ПР' # Обозначение для предмета без типа пары, но с подгруппой
default_date = timey_wimey # Дата по умолчанию = период
default_cab = 'АКТ_зал'
#default_prep = 'Тот-Чьё-Имя-Нельзя-Называть'  # Препод по умолчанию
default_prep = 'Джон Киль'

""" Процедуры форматирования элементов записи """
# Форматирование преподов
def format_prep(prepod):
    repair_pat1 = r'^([а-я. \d]+)\s'             # Должность
    repair_pat2 = r'([А-Я][а-я]+)\s*([А-Я.]{4})' # Препод без должности
    # Для стабильности шаблона, лучше добавлять точку после второго инициала (но ведь лучше сразу адаптировать)
    if not re.search(r'\.$', prepod, flags = re.M):
        prepod += '.'
    prep = ['', re.search(repair_pat2, prepod)[1], re.search(repair_pat2, prepod)[2]]
    if re.search(repair_pat1, prepod, flags = re.M):
        prep[0] = re.search(repair_pat1, prepod, flags = re.M)[1]
        prep[0] = re.sub(r'\s+', ' ', prep[0])
        prep[0] = re.sub(r'\.\s+', '.', prep[0])
    else:
        del(prep[0])
    return ' '.join(prep)

# Форматирование типа пары
def format_tip(tip):
    repair_pat = r'(.*?6.*D.*)|(.*?7.*Z.*)|(.*?т.*р.*я.*)|(.*?л.*к.*я.*)|(.*?а.*б.*)|(.*?п.*а.*к.*)'
    for i, sovp in enumerate(re.findall(repair_pat, tip)[0]):
        if sovp:
            return tip_list[i]

# Форматирование типа подгруппы
def format_group(group):
    repair_pat = r'(\d)'
    return re.findall(repair_pat, group)[0] + 'п/гр'

# Абсолютное разбитие даты
def expand_dates(dates, year, day):
    dates = '; '.join(dates)
    any_dates = r'(?:с\s*[\d.]{5,8}[г.]*\s*по\s*[\d.]{5,8}[г.]*)|(?:[\d,.]+[.\s]*?[\d]{2})'
    dates = re.findall(any_dates, dates)
    repair_pat1 = r'с\s*([\d.]{5,8})[г.]*\s*по\s*([\d.]{5,8})[г.]*' # Дата вида "с..по.."
    repair_pat2 = r'([\d,.]+)[.\s]*([\d]{2})' # Дата вида "день, день,...,месяц
    all_dates = []
    for date in dates:
        if re.search(repair_pat1, date): # Пример входных данных: 'с 13.01 по 06.06'
            start_end = re.search(repair_pat1, date).groups()
            dt_start = list(map(int, start_end[0].split('.')))
            dt_final = list(map(int, start_end[1].split('.')))
            dt_start = dt_date(int(year), dt_start[1], dt_start[0])
            sdvig = 7 # Сдвиг - неделя при любом дне
            while dt_start.weekday() != day:
                dt_start += dt_tdelta(days=1)
            dt_final = dt_date(int(year), dt_final[1], dt_final[0])
            while dt_start <= dt_final:
                all_dates.append(dt_start)
                dt_start += dt_tdelta(days = sdvig)
        else: # Пример входных данных: '03,17,24,31.03'
            date = re.findall(repair_pat2, date)[0]
            data = re.sub(r'\.$', '', date[0], flags = re.M)
            # Если перед месяцем запятая вместо точки
            if int(date[1]) > 12:
                days = re.split(r'[,.]', data[:-3])
                month = data[-2:]
            else:
                days = re.split(r'[,.]', data)
                month = date[1]
            all_dates.extend(dt_date(int(year), int(month), int(day)) for day in days if day)
    # После абсолютного разбития, возвращается отсортированный список дат записи без повторов
    return sorted(list(set(all_dates)))

""" Регулярные шаблоны для переработки информации """
pat1 = r'([А-ЯЁA-Z][а-яaёa-zА-ЯЁA-Z, ()-]{3,})(?:[:;\n\d]|$)' # Отлов предмета (зачёты и дифы исключаются до отлова)
pat2 = r'(?:[а-яёa-z]{2,}[а-яёa-z. \d]+\s{1,})?[А-Я][а-я]+\s*[А-Я.]{3,4}' # Отлов препода и должности (если есть)
pat3 = r'(?:(?:с\s*(?:\d{2}[.,г;\s]*\s*)+по\s*(?:\d{2}(?:[.,г;\s]|$)\s*)+))|(?:(?:\d{2}(?:[.,г;\s]|$)\s*)+)' # Даты
pat4 = r'\d\s*[п]?\s*/\s*гр' # Отлов подгрупп
pat5 = r'(?:[практика]{8})|(?:[лаб.раб ]{8})|(?:[лекция]{6})|(?:[теория]{6})|6D6D6D6|7Z7Z7Z7' # Отлов типа пары

# Разделение по преподам
pattern1a = r'(?:.*?[А-ЯЁA-Z][а-яaёa-zА-ЯЁA-Z, ()-]{3,}(?:[:;\n\d]|$)(?:(?:.(?![А-ЯЁA-Z][а-яaёa-zА-ЯЁA-Z, ()-]{3,}(?:[:;\n\d]|$)))*[А-Я][а-я]+\s*[А-Я.]{3,4}))|(?:.*?[А-ЯЁA-Z][а-яaёa-zА-ЯЁA-Z, ()-]{3,}(?:[:;\n\d]|$))'
# Разделение по предметам
pattern1b = r'(?:.*?(?:[А-ЯЁA-Z][а-яёa-zА-ЯЁA-Z, ()-]{3,}(?:[:;\n\d]|$)).*?)(?=(?:[А-ЯЁA-Z][а-яёa-zА-ЯЁA-Z, ()-]{3,}[:;\n\d])|(?:$))'
# Разделение по преподам
pattern2a = r'.*?(?:[а-яёa-z.]{2,}\s[А-ЯЁA-Z][а-яёa-z\s]+(?:[А-ЯЁA-Z][.]?){2})'
# Разделение по преподам: выделение подгрупп и преподов (если после инфы просто перечисление подгрупп-преподов)
pattern2b = r'(\d\s*[п]?\s*/\s*гр).*?((?:[а-яёa-z]{2,}[а-яёa-z. \d]+\s{1,})?[А-ЯЁA-Z][а-яёa-z ]{2,}(?:[А-ЯЁA-Z][.]?){2})'
# Отлов даты в конце строки (для конкретных случаев, когда косяк в захвате даты)
pattern3 = r'(?:(?:(?:с\s*(?:\d{2}[.,г;\s]*\s*)+по\s*(?:\d{2}[.,г;\s]{0,2}\s*)+))|(?:(?:\d{2}[.,г;\s]{0,2}\s*)+))$'
# Отлов дат, типов пары и групп как [даты, типы пары, группы]
pattern4 = r'((?:(?:с\s*(?:\d{2}[.,г;\s]*\s*)+по\s*(?:\d{2}(?:[.,г;\s]|$)\s*)+))|(?:(?:\d{2}(?:[.,г;\s]|$)\s*)+))|((?:[практика]{8})|(?:[лаб.раб ]{8})|(?:[лекция]{6})|(?:[теория]{6})|6D6D6D6|7Z7Z7Z7)|(\d\s*[п]?\s*/\s*гр)'
# Для лингв. проверки случаев с несколькими преподами
pattern5 = r'.*(?:[А-ЯЁA-Z][а-яaёa-zА-ЯЁA-Z, ()-]{3,})(?:[:;\n\d]|$).*?(?=(?:\d\s*[п]?\s*/\s*гр)|(?:[а-яёa-z]{2,}[а-яёa-z. \d]+\s{1,}[А-ЯЁA-Z][а-яёa-z ]+(?:[А-ЯЁA-Z][.]?){2}))'

""" Преобразование базы обработки в базу разбора (parse)
Здесь данные заботливо перерабатываются в примитивную БД (список списков)
"""
parse = []
day_num = set() # Множество итерации дней
for record in table:
    cabs = iter(record[-1])
    day_num.add(record[0]) # Множество пройденных дней (для определения текущего)
    if not record[2]: # Если инфы в ячейке предмета нет, то запись не обрабатывается
        continue
    """
    print('='*188)
    print(record[0], record[1], record[-1])
    #print(record[2])
    print('-'*188)
    """
    # Первичное обрубание: по преподам (перед предметом может идти инфа о нём...
    # ...но после последнего препода в логическом наборе предмета, начинается инфа о след. предмете
    divide = list(a for a in re.findall(pattern1a, record[2]))
    # Из-за особенностей работы с индексами, обработка оптимальна через while (т.к длина divide может меняться)
    # Минусы: нужно внимательно отлавливать бесконечный цикл
    # Плюсы: для полной обработки не нужно повторно проходить по разбору (уменьшение временной сложности)
    ind = 0 # Абсолютный индекс актуального элемента в divide
    while True:
        rec = divide[ind]
        # Можно обойтись и без условия, но оно помогает избежать лишних проверок (нужно для оптимизации)
        # Цикл для разделения хлама вида "ВМ...МО...Кондратьев В.П." или "Ивент...нормальный предмет и его препод"
        if len(re.findall(pat1, rec)) > 1:
            prepod = re.search(pat2, rec)[0] # Для комфортного добавления препода в строку
            i = 0 # Относительный индекс для отслеживания новых записей
            for predm_rec in re.findall(pattern1b, rec):
                # Выделить предмет в отдельную запись
                divide = divide[:ind+i+1] + [predm_rec] + divide[ind+i+1:]
                i += 1 # Увеличить относительный индекс
                # Ивентовый предмет без препода - нечто вида "09.01.20; Час куратора" (типа пары никогда нет)
                # Если предмет не ивентовый (т.е есть тип пары), а препода нет - "и дайте этому предмету препода"
                if re.search(pat5, predm_rec) and not re.search(pat2, predm_rec):
                    divide[ind+i] += '; '+prepod
            # Удалить изначальную разделяемую запись
            del(divide[ind])
            rec = divide[ind]

        # Формирование разборного списка записей "предмет и инфа о нём"
        # В разборном списке один элемент по умолчанию, и 2+ если запись имеет вид "Предмет...преподы"
        razbor = []
        # Если есть конструкция "предмет...препод1...преподN"
        if len(re.findall(pattern2a, divide[ind])) > 1:
            # Если даты и типы пар общие для подгрупп (т.е остаток = "гр1-препод1, гр2-препод2...")
            if not re.search(pat3, re.sub(pattern5, '', divide[ind])):
                part1 = re.findall(pattern5, divide[ind])[0] # Предмет; даты. - тип пары:
                part2 = re.sub(pattern5, '', divide[ind])    # 1п/гр.: препод1;...; 'Nп/гр.: преподN
                for pp in re.findall(pattern2b, part2):
                    # Так как инфа в записи общая для ячейки, и меняются только препод с подгруппой...
                    # То для каждого типа пары нужно добавить подгруппу (она указывается один раз, из-за чего могут быть баги)
                    # При этом, оригинальное указание подгруппы затирается
                    razbor.append('; '.join([re.sub(pat5, lambda m: m[0]+': '+pp[0], re.sub(pat4, '', part1)), pp[1]]))
            else:
                # Шаг 1) Вырезать предмет (перед ним может быть дата, если первый тип - лекция/практика)
                predmet = re.search(pat1, divide[ind], flags = re.M)[1]
                for_sep = re.sub(predmet, '', divide[ind])
                # Шаг 2) Строка без предметов логически делится по преподам
                for razd in re.findall(pattern2a, for_sep):
                    razbor.append('; '.join([predmet, razd]))
        else:
            razbor.append(divide[ind])
        # День недели - номер пары - предмет - препод - тип пары - подгруппа - даты
        for nabor in razbor:
            # Вырезать предмет из набора в отдельную переменную (если предмет указан)
            if re.search(pat1, nabor, flags = re.M):
                predmet = re.search(pat1, nabor, flags = re.M)[1]
                nabor = re.sub(pat1, '', nabor)
            else: # Если предмета нет (такого быть не должно, но мало ли)
                predmet = 'ОШИБКА, ПРЕДМЕТ НЕ НАЙДЕН'

            # Выделить наборы "даты - тип пары - группа"
            it_inf = iter([a for a in re.findall(pattern4, nabor, flags = re.M) if a]) # Типологический итератор
            dtg = [[[], [], []]] # Логический список дат/типов/групп конструкции
            pred = 9 # Индекс предыдущей группы совпадений, нужен для остановки после комбо + фулсета
            # Комбо, когда 
            for trash in it_inf:
                # Определить, что отловилось (даты - группа №0, типы - 1, группы - 2)
                for i, unint in enumerate(trash):
                    if unint: # [] = False, т.е если не False - элемент = улов
                        # Если не хватает подгруппы, а актуал не подгруппа, то это уже следующий набор
                        if i!=2 and dtg[-1][0] and dtg[-1][1] and not dtg[-1][2]:
                            dtg[-1][2].append(default_grup)
                        if not dtg[-1].count([]) and i!= pred or dtg[-1][1] and i==1: # Если фулсет и не комбо, или следующий набор
                            dtg.append([[], [], []]) # То перейти к следующему элементу
                        dtg[-1][i].append(unint)
                        break # Нет смысла проверять остальное, если искомое нашлось
                pred = i
            # Последняя запись не проверяется на подгруппы (из-за особенностей итерирования)
            if not dtg[-1][2]: # Если нет подгруппы
                dtg[-1][2].append(default_grup) # Значение по умолчанию

            # Исправление возможных ошибок в изначальном расписании
            i = 0 # Опять же, длина списка может меняться, из-за чего приходится использовать while
            while True:
                # Разделение записи по подгруппам, если оно возможно
                if len(dtg[i][2]) > 1:
                    j = len(dtg[i][2])
                    for gr in dtg[i][2]:
                        dtg = dtg[:i+1] + [[dtg[i][0], dtg[i][1], [gr]]] + dtg[i+1:]
                    del(dtg[i])
                    i += j-1
                # Исправление разделения записей вида "ФЗК;  практика: лаб.раб.: 15.04.20г. - 1п/гр"
                if not dtg[i][0]:
                    if i > 0: # Если косяк в "не первой" записи, то склеить её с предыдущей
                        dtg[i-1][1][0] += ', '+dtg[i][1][0] # Склейка типов пары
                        del(dtg[i])
                        i -= 1 # На случай, если дальше что-то будет
                    elif i+1 != len(dtg): # Если косяк в первой записи, то склеить её со следующей
                        dtg[i+1][1][0] += ', '+dtg[i][1][0] # Склейка типов пары
                        del(dtg[i])
                        i -= 1 # На случай, если дальше что-то будет
                # Если запись не содержит тип пары, то он либо был упомянут ранее, либо таков замысел
                if i>0 and not dtg[i][1]:
                    dtg[i][1] = dtg[i-1][1] # Есть предыдущая запись? Взять тип пары из неё
                elif not dtg[i][1]:
                    dtg[i][1].append(default_type_a) # Запись первая? Тип по умолчанию
                # Если все записи были обработаны, исправление ошибок завершается. Иначе, следующая запись
                if i+2 > len(dtg):
                    break
                else:
                    i += 1

            # Форматирование дат, типа пары и подгруппы
            for f in range(len(dtg)):
                dtg[f][0] = expand_dates(dtg[f][0], year, len(day_num)-1)
                if dtg[f][1][0] != default_type_a:
                    dtg[f][1] = [format_tip(dtg[f][1][0])]
                if dtg[f][2][0] != default_grup:
                    dtg[f][2] = [format_group(dtg[f][2][0])]
            
            # Выделить препода (выделяется после выделения типов пар)
            # Если выделить до, то при ошибке в синтаксисе изначального расписания можно поймать:
            # Препод = "теория ст.преподаватель БелкинаА.В."
            if re.findall(pat2, nabor): # Если есть препод, отловить его с должностью (если указана)
                prepod = [format_prep(re.findall(pat2, nabor)[0])]
            else: # Если препод не указан
                if re.findall(pat5, razbor[i], flags = re.I): # Если у текущей пары есть тип, то она не "особая"
                    prepod = [format_prep(parse[-1][3])]
                else: # Пара особая = мероприятие, час куратора и т.п = препод не указывается
                    prepod = [default_prep]

            # Для каждого набора "тип пары - подгруппа - даты" создать запись в БД
            for info in dtg:
                # Финальная корректировка типа
                if len(parse) and info[1][0]==default_type_a:
                    if parse[-1][2] == predmet: # Если предмет как в предыдущем, а тип базовый, то ошибка в типе
                        info[1] = [parse[-1][4]]
                    elif info[2][0] != default_grup: # Если тип базовый, но есть подгруппы - сменить на базовый №2
                        info[1][0] = default_type_b
                #"""
                # Расстановка кабинетов
                # Если сейчас идёт особый предмет
                if info[1][0] == default_type_a:
                    cab = default_cab
                # Если кабинеты есть, то подбирать в зависимости от условий
                elif record[-1]:
                    if not len(parse) or record[1]!=parse[-1][1]: # Если номер пары изменился
                        try:
                            cab = next(cabs)
                        except:
                            cab = parse[-1][7]
                    else: # Если это всё тот же набор кабинетов (для каждого номера пары набор один)
                        # Если тип пары лекция, и до этого тоже была лекция:
                        if info[1][0]==tip_list[2] and parse[-1][4]==tip_list[2]:
                            cab = parse[-1][7] # Взять предыдущий кабинет
                        # Если кабинетов больше чем преподов, и предыдущий тип - лекция
                        elif parse[-1][4]==tip_list[2] and len(record[-1])>len(re.findall(pat2, record[2])):
                            try:
                                cab = next(cabs)
                            except:
                                cab = parse[-1][7]                            
                        else:
                            # Если препод не изменился:
                            if parse[-1][3] == prepod[0]:
                                cab = parse[-1][7] # Взять предыдущий кабинет
                            else:
                                try:
                                    cab = next(cabs)
                                except:
                                    cab = parse[-1][7]
                # Если кабинеты не указаны, то значение по умолчанию
                else:
                    cab = default_cab
                #"""
                # Занесение в базу
                parse.append([record[0],   # День недели  # 0
                              record[1],   # Номер пары   # 1
                              predmet,     # Предмет      # 2
                              prepod[0],   # Препод       # 3
                              info[1][0],  # Тип пары     # 4
                              info[2][0],  # Подгруппа    # 5
                              info[0],     # Даты         # 6
                              cab          # Кабинет      # 7
                              ])
                #print(parse[-1][:-2])
                #print(parse[-1][:-2], [parse[-1][-1]])
                #print(parse[-1][-2])
                # Проверка на разброс одной логической записи по двум записям (когда криво записано в таблице)
                # "Информатика; 01,22,29.03; 05,12.04.19г. - 2п/гр.; лаб.раб: 19,26.04; 10,17,24.05.19г. -2п/гр.: доцент ОбвинцевО.А.;"
                if len(parse)>1 and parse[-2][:-2]==parse[-1][:-2]:
                    parse[-2][-2].extend(parse[-2][-2])
                    del(parse[-1])
                    # Удаление повторов дат
                    parse[-1][6] = sorted(list(set(parse[-1][6])))
        
        if ind+1 == len(divide):
            break
        else:
            ind += 1

"""
for i, record in enumerate(parse):
    if i > 0:
        if parse[i-1][1] != parse[i][1]:
            print()
    if record[7][0] == 'N':
        print(record[7][1:])
        print(record[:6])
"""

# Тестовый вывод для сверки. Без дат - они занимают много места
"""
parse_title = 'Расписание '+str(title[v])+' на '+year+'-й год. Учебная часть семестра идёт с '+period[0]+' по '+period[1]
print(f"\n{parse_title : ^188}\n")
print('='*187)
print(f"| {'День' : ^4} | {'№' : ^1} | {'Предмет' : ^66} | {'Препод' : ^57} | {'Тип' : ^15} | {'Для кого' : ^11} | {'Каб' : ^11} |")
print('-'*187)
for i, record in enumerate(parse):
    if i>0 and record[0]!=parse[i-1][0]:
        print('-'*187)
    print(f"| {record[0] : ^4} | {record[1] : ^1} | {record[2] : ^66} | {record[3] : ^57} | {record[4] : ^15} | {record[5] : ^11} | {record[7] : ^11} |")
    print(f'{record[6]}')
print('='*187)
"""

""" Анализ (для вывода по подгруппам)
1) Создаётся словарь с инфой о предметах
Его структура такова:
- Ключ  предмет
- Значение - список наборов "тип пары - подгруппа"
2) Словарь анализируется по количеству подгрупп, и конвертится в нечто по типу:
   - Ключ: предмет
   - Значение: список, где на 1-м месте кол-во подгрупп типа пары, а на 2-м - его замена при ошибках
"""

# Составление основы начальной базы анализа подгрупп
predmets = {}
for record in parse:
    if record[2] not in predmets:
        predmets[record[2]] = [[record[4], record[5]]]
    else:
        if [record[4], record[5]] not in predmets[record[2]]:
            predmets[record[2]].append([record[4], record[5]])

# Переработка базы анализа подгрупп
for predmet in sorted(predmets.keys()):
    good, bad, temp = {}, {}, {}
    for info in sorted(predmets[predmet]):
        if info[0] not in temp:
            temp[info[0]] = [info[1]]
        else:
            temp[info[0]].append(info[1])
    # Прогон каждого набора 'Тип пары': [подгруппы]
    for info in temp.keys():
        # Если подгруппа - общее умолчание, то всё хорошо
        if temp[info] == [default_grup]:
            good[info] = temp[info]
        # Если подгруппы есть, то среди них могут быть пропуски
        # Пропуски возникают, когда в расписании лабу случайно обозвали практикой (и т.п)
        elif len(temp[info]) > 1:
            zbs, mx = True, 0 # Наличие пропусков, максимальная подгруппа (для пропусков)
            # Прогон по всем подгруппам
            for g_ind, grp in enumerate(temp[info], start = 1):
                # Если где-то номер подгруппы не совпадает с тем что должен быть, всё плохо
                if grp != default_grup and g_ind != int(grp[0]):
                    if info not in bad:
                        bad[info] = []
                    bad[info].append(grp)
                    zbs = False
            if zbs == True:
                good[info] = temp[info]
        # Если есть подгруппы, но в списке всего одна - что-то пошло не так
        else:
            bad[info] = [temp[info][0]]
    # Если есть как хорошие, так и плохие типы пар
    if bad and good:
        # Прогон каждого плохого типа пары
        for b in bad:
            maybe = {k: v for k, v in good.items() if v != [default_grup]} # Типы пар, которые могут юзаться как заменитель
            # Если есть ровно один заменитель, добавить к нему возможные подгруппы. Занести заменитель как значение плохиша
            if len(maybe) == 1:
                for val in bad[b]:
                    if val not in good[list(maybe.keys())[0]]:
                        good[list(maybe.keys())[0]].append(val)
                bad[b] = list(maybe.keys())[0]
            # Если заменителя в хорошем нет, то мб он появится в плохом
            elif not len(maybe):
                good[b] = bad[b]
            # Если потенциальных заменителей несколько, выбрать тот что ближе соответствует подгруппам
            else:
                mx = max(list(map(int, (a[0] for a in bad[b])))) # Выделить максимальную подгруппу в плохише
                zbs, zam = False, False
                for mb in maybe: # Точный подбор заменителя
                    for g in maybe[mb]: # Проверка всех подгрупп
                        if int(g[0]) == mx:
                            zam, zbs = mb, True
                            break
                if not zbs:
                    for mb in maybe: # Примерный подбор заменителя (если точный не помог)
                        if mx in range(int(g[0])-1, int(g[0])+2):
                            zam, zbs = mb, True
                            break
                good[zam].extend(b.items()[0][1])
                for val in bad[b]:
                    if val not in good[zam]:
                        good[zam].append(val)
                bad[b] = zam
    # Если хорошего списка нет, то попалось кривое исключение
    elif bad:
        good = bad
        bad = {}
    # После обработки ошибок, нужно переработать temp как 'Ключ': [Кол-во подгрупп, заменитель]
    for t in temp:
        # Фикс лишнего добавления. Делается здесь, чтобы избежать ошибок итерирования по словарю bad
        if t in bad and t in good:
            bad.pop(t)
        # Если ключа нет в плохих, или он есть и там и там, то всё ок
        if t in [list(bad.keys()), list(good.keys())]:
            temp[t] = [len(good[t]), t]
        # Если ключа нет в хороших, то взять кол-во подгрупп из плохих, заменитель из него же
        elif t in list(bad.keys()):
            temp[t] = [len(temp[t]), bad[t]]
        # Если ключа нет в плохих, то заменитель не нужен => вместо него ставится ключ
        else:
            temp[t] = [len(good[t]), t]
    predmets[predmet] = temp

""" Запись считанной базы в таблицу """
# Так как openpyxl при обращении к ячейке сразу хранит её, обработка ячейки удобна через _
wb = Workbook() # Создание новой таблицы

""" Константные стили и названия """
# Стиль шапки
st_title = NamedStyle(name = 'Шапка')
st_title.font = Font(name = 'Book Antiqua', size = 14)
st_title.alignment = Alignment(horizontal = 'center', vertical = 'center')
st_title.border = Border(left = Side(border_style = 'medium'),  right = Side(border_style = 'medium'),
                         top  = Side(border_style = 'medium'), bottom = Side(border_style = 'medium'))

# Стиль дней
st_days = NamedStyle(name = 'Дни')
st_days.font = Font(name = 'Bookman Old Style', size = 14, bold = True)
st_days.alignment = Alignment(horizontal = 'center', vertical = 'center', textRotation = 90)
st_days.border = Border(left = Side(border_style = 'medium'), right  = Side(border_style = 'medium'),
                        top  = Side(border_style = 'medium'), bottom = Side(border_style = 'medium'))

# Стиль базовой ячейки
st_baze = NamedStyle(name = 'Базовая ячейка')
st_baze.font = Font(name = 'Plantagenet Cherokee', size = 14)
st_baze.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
st_baze.border = Border(left = Side(border_style = 'thin'), right  = Side(border_style = 'thin'),
                        top  = Side(border_style = 'thin'), bottom = Side(border_style = 'thin'))

# Стиль для номеров пары и времени
st_razn = NamedStyle(name = 'Для инфополей')
st_razn.font = Font(name = 'Plantagenet Cherokee', size = 14)
st_razn.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
st_razn.border = Border(left = Side(border_style = 'medium'), right  = Side(border_style = 'medium'),
                        top  = Side(border_style = 'thin'),   bottom = Side(border_style = 'thin'))

# Стиль пустой ячейки
st_null = NamedStyle(name = 'Круговерть пустоты')
st_null.border = Border(left = Side(border_style = None), right  = Side(border_style = None),
                        top  = Side(border_style = None), bottom = Side(border_style = None))
st_null.fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')

# Названия столбцов
parse_title = ['Дни', '№ пары', 'Время', 'Ауд', 'Преподаватель']

# Полные названия месяцев
mos = {1: 'Январь',    2: 'Февраль',  3: 'Март',    4: 'Апрель',
       5: 'Май',       6: 'Июнь',     7: 'Июль',    8: 'Август',
       9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'}

# Сокращения месяцев (альтернатива "mos[date[0].month][0].lower()+mos[date[0].month][1:3]")
# Улучшает читаемость кода + повышает кастомизацию
mos_s = {1: 'янв',   2: 'фев',  3: 'марта', 4: 'апр',
         5: 'мая',   6: 'июня', 7: 'июля',  8: 'авг',
         9: 'сент', 10: 'окт', 11: 'нояб', 12: 'дек'}

# Расписание звонков в будние дни. 5-минутные перерывы не учитываются, т.к преподы их не делают
time_budni = {0: '08:30 - 10:05', 1: '10:15 - 11:50', 2: '12:35 - 14:10',
              3: '14:20 - 15:55', 4: '16:05 - 17:40', 5: '17:50 - 19:25'}

# Расписание звонков в субботу. 5-минутные перерывы не учитываются, т.к преподы их не делают
time_vihod = {0: '08:30 - 10:05', 1: '10:15 - 11:50', 2: '12:00 - 13:40',
              3: '13:50 - 15:25', 4: '15:35 - 17:10', 5: '17:20 - 18:55'}

""" Определение периода для дней в шапке + ориентации при записи инфы о предмете """
tdt = [] # Список для хранения границ всех учебных недель (элемент - кортеж с границами одной недели)

dt_start = list(map(int, period[0].split('.')))
dt_final = list(map(int, period[1].split('.')))
dt_start = dt_date(int(year), dt_start[1], dt_start[0])
dt_final = dt_date(int(year), dt_final[1], dt_final[0])

s = 0 # Начало учебной недели
while dt_start <= dt_final:
    # Если добавляется первая или новая неделя
    if s==0 or dt_start.weekday()==0:
        s = dt_start
    # Если уже суббота, нужно пропустить воскресенье
    elif dt_start.weekday() == 5:
        tdt.append([s, dt_start])
        dt_start += dt_tdelta(days = 1)
    dt_start += dt_tdelta(days = 1)

# Возврат начального значения первой недели расписания
dt_start = list(map(int, period[0].split('.')))
dt_start = dt_date(int(year), dt_start[1], dt_start[0])

# Заполнение листов таблицы и их шапок
for d_ind, date in enumerate(tdt):
    # Учебный месяц изменяется при определённом наборе условий
    if any([all([d_ind,
                 any([tdt[d_ind-1][0].month != date[0].month,
                      tdt[d_ind-1][1].month != date[1].month]),
                 any([int(date[0].day) in range(1, 7),
                      int(date[1].day) in range(1, 7)]),
                 mos[date[0].month] not in wb.sheetnames]),
            not d_ind]):
        # Если в месяце была одна неделя (при смене месяца), то нужно просто объединить месяцы
        if not wb.sheetnames[0] == 'Sheet' and ws.max_column < 5:
            ws.title = mos[date[0].month]
        # Если в следующем месяце будет всего одна неделя (на случай чумы, апокалипсиса или сессии), то ничего не менять
        elif not(d_ind+1 == len(tdt) or d_ind and tdt[d_ind-1][0].month != date[0].month and date[0].month != tdt[d_ind+1][0].month):
            # Если добавляется первый лист, то можно просто переименовать стандартный
            if wb.sheetnames[0] == 'Sheet':
                ws = wb.active
                ws.title = mos[date[0].month]
            # Если обычный новый учебный месяц, то создать новый лист
            else:
                wb.create_sheet(mos[date[0].month])
            # Сделать актуальным листом последний лист
            ws = wb.worksheets[-1]
            # Заполнение первых трёх столбцов шапки
            _ = ws.cell(column=ws.max_column,   row=1, value=parse_title[0]).style = st_title
            _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[1]).style = st_title
            _ = ws.cell(column=ws.max_column+1, row=1, value=parse_title[2]).style = st_title
    # Запись учебной недели в красивом формате
    m1, m2 = mos_s[date[0].month], mos_s[date[1].month]
    val = date[0].strftime('%d') + ' ' + m1 + ' - ' + date[1].strftime('%d') + ' ' + m2
    _ = ws.cell(column=ws.max_column+1, row=1, value=val).style = st_title

    # Если последняя учебная неделя, или произошла смена листа, то доставить титулы кабинетов и преподов
    if d_ind == len(tdt)-1 or ws.max_column == 4 and len(wb.worksheets) > 1:
        if d_ind == len(tdt)-1:
            pred = wb.worksheets[-1]
        else:
            pred = wb.worksheets[-2]
        _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[3]).style = st_title
        _ = pred.cell(column=pred.max_column+1, row=1, value=parse_title[4]).style = st_title   
  


""" Заполнение основной части таблицы """
# Константные установки и основные индексы
grp2, grp3 = 0, 0 # Подгруппы в разделении на 2 подгруппы, и в разделении на 3 подгруппы
grps = {1: 0, 2: grp2, 3: grp3} # Для выбора подгруппы
c_cabs = True # Обрезать кабинет?
c_prep = True # Обрезать препода?
c_grps = True # Обрезать подгруппу?

days = iter(days_names) # Итератор по дням
act_r = 1  # Индекс актуальной строки в таблицы
act_bd = 0 # Актуальная запись в базе разбора

# Базовое заполнение таблицы предметами
# Прогон по номерам пары в шаблоне, чтобы день не мог получиться таким как "№2, №4, №6"
for num in range(len(time_budni) * 6):
    # Если следующий день
    if num%6 < (num-1)%6:
        day = days_names[next(days)]
    # Если для дня пары закончились, но время ещё есть - лучше просто перейти к следующему дню
    if act_bd == len(parse) or act_bd and parse[act_bd][0] != day:
        continue
    act_r += 1

    # Занести день, номер и время пары на все листы. Применить "разный" стиль к кабинетам и преподам, мало ли где пропуск
    for ws in wb:
        _ = ws.cell(column = 1, row = act_r, value = day).style = st_days
        _ = ws.cell(column = 2, row = act_r, value = num%6 + 1).style = st_razn
        if day != days_names[5]:
            _ = ws.cell(column = 3, row = act_r, value = time_budni[num % 6]).style = st_razn
        else:
            _ = ws.cell(column = 3, row = act_r, value = time_vihod[num % 6]).style = st_razn
        ws.cell(column = ws.max_column-1, row = act_r).style = st_razn
        ws.cell(column = ws.max_column,   row = act_r).style = st_razn

    # Проверить соответствие записи в БД актуальному дню и номеру пары
    if not act_bd or parse[act_bd][0] == day and parse[act_bd][1] == num%6+1:
        # Если соответствует, то обработать все записи для актуального дня и номера пары
        while True:
            # Итераторы для записи инфы о предмете
            weeks = iter(tdt)               # Итератор по учебным неделям семестра
            it_week = next(weeks)           # Актуал для проверяемой недели
            it_sheets = iter(wb.sheetnames) # Итератор по листам книги
            ws = wb[next(it_sheets)]        # Актуальный лист
            act_c = 4                       # Актуальный столбец. Для проверки на смену листа
            rec = parse[act_bd]             # Актуальная запись в базе разбора

            # Проверить подгруппу у типа пары в графе замены
            zam = predmets[rec[2]][rec[4]][1]   # На что заменить тип пары
            gr = grps[predmets[rec[2]][zam][0]] # Выбранная подгруппа для типа пары

            # Если на подгруппу пофиг (0), или подгруппа соответствует текущей, то инициировать занесение информации в таблицу
            if not gr or gr == int(rec[5][0]):
                # Прогон по каждой дате в записи
                for date in rec[6]:
                    # Если дата не входит в период расписания
                    if date < dt_start or date > dt_final:
                        continue
                    # Если записи нет на этой учебной неделе: поиск нужной недели, т.е...
                    while not (it_week[0] <= date <= it_week[1]):
                        it_week = next(weeks) #...переход к новой неделе
                        act_c += 1  #...переход на следующий столбец таблицы
                        # Проверить, есть ли новая учебная неделя на текущем листе: если актуал столбца на кабинетном столбце, то...
                        if act_c > ws.max_column-2:
                            ws = wb[next(it_sheets)] #...переход на следующий лист
                            act_c = 4                #...обновление актуального столбца

                    # Добавление инфы о предмете и подгруппе
                    if c_grps and rec[5] != default_grup: # Если нужно обрезать подгруппу
                        rec[5] = rec[5][0]
                    pat_rec = f'{rec[2]}: {rec[4]} ({rec[5]})' # Форматная запись предмета
                    if ws.cell(column=act_c, row=act_r).value:
                        val = '\n'.join([ws.cell(column=act_c, row=act_r).value, pat_rec])
                    else:
                        val = pat_rec
                    _ = ws.cell(column=act_c, row=act_r, value=val).style = st_baze

                    # Добавление инфы о кабинете и преподе
                    if c_cabs: # Если нужно обрезать кабинет
                        rec[7] = re.sub(r'\s*УК№\d', '', rec[7])
                    if c_prep and rec[3] != default_prep: # Если нужно обрезать должность препода
                        rec[3] = re.search(r'[А-Я][а-я]+\s*[А-Я.]{4}', rec[3])[0]
                    rzd_a = re.split(r'[\n,] ?', str(ws.cell(column=ws.max_column-1, row=act_r).value))
                    rzd_b = re.split(r'[\n,] ?', str(ws.cell(column=ws.max_column,   row=act_r).value))
                    if not ws.cell(column=ws.max_column, row=act_r).value or rzd_b[-1]!=rec[3] or rzd_b[-1]==rec[3] and rzd_a[-1]!=rec[7]:
                        # Добавление инфы о кабинете
                        # Случай 1) Пустая ячейка
                        if not ws.cell(column=ws.max_column-1, row=act_r).value:
                            val = rec[7]
                        # Случай 2) Препод не изменился, но кабинет другой
                        elif rzd_b[-1] == rec[3]:
                            val = ', '.join([ws.cell(column=ws.max_column-1, row=act_r).value, rec[7]])
                        # Случай 3) Не пустая ячейка
                        else:
                            val = '\n'.join([ws.cell(column=ws.max_column-1, row=act_r).value, rec[7]])
                        _ = ws.cell(column=ws.max_column-1, row=act_r, value=val)

                        # Добавление инфы о преподе
                        # Случай 1) Пустая ячейка
                        if not ws.cell(column=ws.max_column, row=act_r).value:
                            val = rec[3]
                        # Случай 2) Препод не изменился, но кабинет другой
                        elif not(rzd_b[-1] == rec[3]):
                            val = '\n'.join([ws.cell(column=ws.max_column, row=act_r).value, rec[3]])
                        else:
                            continue
                        _ = ws.cell(column=ws.max_column, row=act_r, value=val)

            act_bd += 1
            if act_bd == len(parse) or rec[0]!=parse[act_bd][0] or rec[1]!=parse[act_bd][1]:
                break
    # Форматирование инфы о кабинете (возможно только когда она вся заполнена)
    for ws in wb:
        cab = ws.cell(column = ws.max_column-1, row = act_r).value
        if cab:
            cab = set(cb for cb in re.split(r'[\n,] ?', cab) if cb)
            # Если инфа о кабинетах - один дублирующийся кабинет, то схлопнуть её в него
            if len(cab) == 1:
                cab = list(cab)[0]
                try: # Эксель порой ругается, что "число сохранено как текст". Бесит
                    cab = int(cab)
                except:
                    pass
                ws.cell(column = ws.max_column-1, row = act_r).value = cab


# Процедура для проверки пересечения объединённых ячеек
def cross_merge(cell_1, cell_2, merge_list):
    # Возвращается координатный список для объединения ячейки
    # [start_column, end_column]
    # Во избежание попыток объединения пустых частей в объединённых ячейках
    if not cell_1.value and not cell_2.value:
        return False
    # Определение типов ячеек (модуль не всегда корректно его предоставляет, хз почему, но доказано логами и крашами)
    t1 = t2 = None
    for cell_m in merge_list:
        if not t1 and cell_1.coordinate in cell_m:
            t1, cell_1 = 'MergedCell', re.findall(r'[A-Z](?=\d)', cell_m.coord)
        if not t2 and cell_2.coordinate in cell_m:
            t2, cell_2 = 'MergedCell', re.findall(r'[A-Z](?=\d)', cell_m.coord)
    # После определения, всегда есть четыре варианта ситуаций
    # Вариант 1) Обе ячейки - обычные. То есть, наложения между ними быть не может
    if not t1 and not t2:
        return [cell_1.coordinate[0], cell_2.coordinate[0]]
    # Группа случаев с объединёнными ячейками. Здесь нужно проверять равенство пограничных столбцов
    else:
        # Вариант 2) Обе ячейки - объединённые
        if t1 == t2:
            # Всё збс только если равенство столбцов
            if cell_1 == cell_2:
                return [cell_1[0], cell_1[1]]
            else:
                return False
        # Вариант 3) Первая ячейка - объединённая, вторая - обычная
        elif t1 and not t2:
            # Всё збс если каким-то фиговым листом объединённая ячейка объединяет всего одну ячейку
            if cell_1[0] == cell_1[1] == cell_2.coordinate[0]:
                return [cell_1[0], cell_1[1]]
            else:
                return False
        # Вариант 4) Первая ячейка - обычная, вторая - объединённая
        else:
            # Аналогично случаю три, но проверяется вторая ячейка
            if cell_2[0] == cell_2[1] == cell_1.coordinate[0]:
                return [cell_2[0], cell_2[1]]
            else:
                return False

# Донастройка визуала таблицы. Несколько прогонов, чтобы не напортачить в данных
for ws in wb:
    # Первый прогон - проверка "Есть ли пустые строки В КОНЦЕ дня"
    # Проверка идёт для всех строк, но без шапки и начиная с конца
    for row in range(ws.max_row, 1, -1):
        # Если нет кабинета и препода, а строка последняя, или у её нижнего соседа другой день:
        if all([not ws.cell(column=ws.max_column-1, row=row).value,
                not ws.cell(column=ws.max_column  , row=row).value,
                row == ws.max_row or ws.cell(column=1, row=row).value!=ws.cell(column=1, row=row+1).value
               ]):
            ws.delete_rows(row)# Удалить строку

    # Второй прогон - выделение пустот
    # Прогон идёт по всем ячейкам с инфой о предметах
    for row in range(2, ws.max_row+1):
        for col in range(4, ws.max_column-1):
            # Если ячейка пустая, применить к ней стиль пустой ячейки
            if not ws.cell(column=col, row=row).value:
                ws.cell(column=col, row=row).style = st_null

    # Третий прогон - проверка "Слева - направо"
    for row in range(2, ws.max_row+1):
        # Пустая строка объединяется сразу
        if not ws.cell(column=4, row=row).value and not ws.cell(column=ws.max_column-1, row=row).value:
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=ws.max_column-2)
            ws.cell(column = 2, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
            ws.cell(column = 3, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
            ws.cell(column = ws.max_column-1, row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
            ws.cell(column = ws.max_column,   row = row).fill = PatternFill(patternType = 'lightDown', start_color = '00ff27')
        else:
            sc = ec = 4
            for col in range(4, ws.max_column-1):
                # Правая ячейка такая же как эта? Сдвинуть правую границу
                if ws.cell(column=col, row=row).value == ws.cell(column=col+1, row=row).value:
                    ec += 1
                # Нет? Вольтрон
                else:
                    if sc != ec:
                        ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=ec)
                    sc = ec = col+1

    # Четвёртый прогон - объединение дней и визуальное выделение дней. Начинается со столбца №2 из-за красивой рамки под временем и номером
    for col in range(2, ws.max_column+1):
        # Сброс нижних границ дня и объединения
        de = oe = ws.max_row
        sc = ec = 4
        # Проверка каждой строки, вплоть до шапки
        for row in range(ws.max_row, 1, -1):
            c1, c2 = ws.cell(column=col, row=row), ws.cell(column=col, row=row-1)
            coords = cross_merge(c1, c2, ws.merged_cells.ranges)
            if coords:
                sc, ec = ws[coords[0]][0].column, ws[coords[1]][0].column
            # Если день не сменился, то можно проверять то что внутри дня
            if ws.cell(column=1, row=row).value == ws.cell(column=1, row=row-1).value:
                # Если верхняя ячейка не подходит для объединения, завершить выделение и начать новое
                if c1.value != c2.value or not coords:
                    if row != oe: # Область выделения должна включать больше одной ячейки
                        # Openpyxl немного туповат и не объединяет объединённые ячейки (даже такие, которые можно объединить)
                        for rw in range(row, oe+1):
                            # Поэтому, всё в объединяемой области нужно разбить
                            try:
                                ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                            except:
                                pass
                        ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)
                    oe = row - 1
            # Если день сменился, то принудительно завершить выделение области + стилизовать день
            else:
                # Объединить имеющуюся область выделения, если в ней нет проблем
                if row != oe:
                    # Openpyxl немного туповат и не объединяет объединённые ячейки (даже такие, которые можно объединить)
                    for rw in range(row, oe+1):
                        # Поэтому, всё в объединяемой области нужно разбить
                        try:
                            ws.unmerge_cells(start_row=rw, start_column=sc, end_row=rw, end_column=ec)
                        except:
                            pass
                    ws.merge_cells(start_row=row, start_column=sc, end_row=oe, end_column=ec)
                # Установить стилизованное выделение для границ дня
                ws.cell(column=col, row=de).border = Border(left   = ws.cell(column=col, row=de).border.left,
                                                            right  = ws.cell(column=col, row=de).border.right,
                                                            top    = ws.cell(column=col, row=de).border.top,
                                                            bottom = Side(border_style = 'medium'))
                # Объединить день, но только когда обход уже на последнем столбце (иначе творится дичь + ненужное выделение)
                if col == ws.max_column:
                    ws.merge_cells(start_row=row, start_column=1, end_row=de, end_column=1)
                # Обновить нижние границы дня и объединения
                de = oe = row-1

            
# Корректировка высоты строки и ширины столбца
for ws in wb:
    for column in ws.columns:
        max_len = max(max((len(r) for r in re.split(r'\n', str(cell.value)))) for cell in column)
        max_len = (lambda m: m*2 if m<10 else m*1.5 if m<20 else m*1.45)(max_len)
        ws.column_dimensions[column[0].column_letter].width = max_len+1
    for row in ws.rows:
        max_str = max(len(re.split(r'\n', str(cell.value))) for cell in row) + 1
        ws.row_dimensions[row[0].row].height = max_str * 17
    ws.freeze_panes = 'D2'

name = f'F:/CoolResp/Резы/Респа для {str(title[v])} на {year} год' + '.xlsx'
wb.save(name)
